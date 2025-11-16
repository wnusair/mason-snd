"""
Tournaments Blueprint

This module handles all tournament-related functionality including creation,
signup management, judge requests, results submission, and partner matching.

Key Features:
    - Tournament Creation: Admins can create tournaments with custom forms
    - Tournament Signup: Students sign up for events within tournaments
    - Judge Requests: Students can request their parents to judge
    - Results Submission: Students and admins can submit tournament performance
    - Partner Matching: Support for partner events (e.g., Public Forum)
    - Custom Forms: Tournaments can have dynamic form fields for additional data

Workflow:
    1. Admin creates tournament with dates and deadlines
    2. Admin optionally adds custom form fields
    3. Students sign up for tournament events
    4. Students request judges (parents) if bringing one
    5. After tournament, students submit results
    6. Admin finalizes and publishes results
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, jsonify

from mason_snd.extensions import db
from mason_snd.models.auth import User, Judges
from mason_snd.models.tournaments import (
    Tournament, Tournament_Performance, Tournaments_Attended,
    Form_Responses, Form_Fields, Tournament_Signups, Tournament_Judges
)
from mason_snd.models.events import User_Event, Event
from mason_snd.utils.race_protection import prevent_race_condition
from mason_snd.utils.auth_helpers import redirect_to_login

from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from io import BytesIO
import pytz

try:
    import pandas as pd
    import openpyxl
except ImportError:
    pd = None
    openpyxl = None

# Timezone constant
EST = pytz.timezone('US/Eastern')

# Blueprint configuration
tournaments_bp = Blueprint('tournaments', __name__, template_folder='templates')


@tournaments_bp.route('/')
def index():
    """
    Display list of tournaments categorized as upcoming or past.
    
    Shows all tournaments sorted by date, with separation between future
    tournaments (upcoming) and tournaments that have already occurred (past).
    
    Returns:
        Rendered template with upcoming_tournaments and past_tournaments lists
    """
    tournaments = Tournament.query.all()

    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        return redirect_to_login("Please log in")

    now = datetime.now(EST)
    
    # Separate tournaments into upcoming and past
    upcoming_tournaments = []
    past_tournaments = []
    
    for tournament in tournaments:
        # Make tournament date timezone-aware if it's naive
        tournament_date = tournament.date
        if tournament_date.tzinfo is None:
            tournament_date = EST.localize(tournament_date)
        
        if tournament_date >= now:
            upcoming_tournaments.append(tournament)
        else:
            past_tournaments.append(tournament)
    
    # Sort tournaments by date
    upcoming_tournaments.sort(key=lambda t: t.date)
    past_tournaments.sort(key=lambda t: t.date, reverse=True)

    return render_template(
        'tournaments/index.html',
        upcoming_tournaments=upcoming_tournaments,
        past_tournaments=past_tournaments,
        user=user,
        now=now
    )



@tournaments_bp.route('/add_tournament', methods=['POST', 'GET'])
@prevent_race_condition(
    'add_tournament',
    min_interval=2.0,
    redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.index'))
)
def add_tournament():
    """
    Create a new tournament (admin only).
    
    GET: Display tournament creation form
    POST: Create tournament with provided details
    
    Form Fields:
        - name: Tournament name
        - address: Tournament location
        - date: Tournament date and time (YYYY-MM-DDTHH:MM)
        - signup_deadline: Deadline for signups (YYYY-MM-DDTHH:MM)
        - performance_deadline: Deadline for results submission (YYYY-MM-DDTHH:MM)
    
    Access: Requires role >= 2 (Admin)
    
    Returns:
        GET: Rendered tournament creation form
        POST: Redirect to tournaments index with success message
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
    
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))
    
    if request.method == "POST":
        name = request.form.get("name")
        address = request.form.get("address")
        date_str = request.form.get("date")  # "YYYY-MM-DDTHH:MM"
        signup_deadline_str = request.form.get("signup_deadline")
        performance_deadline_str = request.form.get("performance_deadline")
        created_at = datetime.now(EST)

        try:
            # Convert string inputs to datetime objects
            date = datetime.strptime(date_str, "%Y-%m-%dT%H:%M")
            signup_deadline = datetime.strptime(signup_deadline_str, "%Y-%m-%dT%H:%M")
            performance_deadline = datetime.strptime(performance_deadline_str, "%Y-%m-%dT%H:%M")
        except ValueError:
            flash("Invalid date format. Please use the date pickers.", "error")
            return render_template("tournaments/add_tournament.html")

        new_tournament = Tournament(
            name=name,
            date=date,
            address=address,
            signup_deadline=signup_deadline,
            performance_deadline=performance_deadline,
            created_at=created_at
        )

        db.session.add(new_tournament)
        db.session.commit()

        users = User.query.all()
        current_tournament = Tournament.query.filter_by(name=name).first()
        events = Event.query.all()
        signup_time = datetime.now(EST)
        for user in users:
            for event in events:
                tournament_signup = Tournament_Signups(
                    user_id = user.id,
                    tournament_id = current_tournament.id,
                    event_id = event.id,
                    created_at = signup_time
                )
                db.session.add(tournament_signup)
        db.session.commit()


    return render_template("tournaments/add_tournament.html")

@tournaments_bp.route('/add_form', methods=['GET', 'POST'])
@prevent_race_condition('add_form', min_interval=1.5, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.index')))
def add_form():
    """
    Add custom form fields to a tournament (admin only).
    
    Allows admins to create dynamic form fields that students must complete
    during tournament signup. Supports multiple field types including text,
    textarea, select dropdowns, and checkboxes.
    
    GET: Display form creation interface with tournament selection
    POST: Create form fields for selected tournament
    
    Form Field Attributes:
        - label: Field label/question text
        - type: Field type ('text', 'textarea', 'select', 'checkbox')
        - options: Comma-separated options for select/checkbox fields
        - required: Boolean indicating if field is mandatory
        - tournament_id: Tournament to attach fields to
    
    Field Types:
        - text: Single-line text input
        - textarea: Multi-line text input
        - select: Dropdown selection
        - checkbox: Yes/No checkbox
    
    Features:
        - Add multiple fields at once (dynamic form builder)
        - Optional vs required field designation
        - Dropdown options for select fields
        - Supports special questions (e.g., "Are you bringing a judge?")
    
    Access: Requires role >= 2 (Admin)
    
    Returns:
        GET: Rendered form creation interface
        POST: Redirect to tournaments index with success message
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
        
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))

    tournaments = Tournament.query.all()

    if request.method == 'POST':
        tournament_id = request.form.get('tournament_id')

        # Get multiple field inputs
        labels = request.form.getlist('label')
        types = request.form.getlist('type')
        options_list = request.form.getlist('options')
        
        # Create a field entry for each input group
        for i in range(len(labels)):
            label = labels[i]
            field_type = types[i]
            options = options_list[i] if i < len(options_list) and options_list[i] != "" else None
            # Check if checkbox for this specific field is checked using unique name
            required = request.form.get(f'required_{i}') == '1'

            new_field = Form_Fields(
                label=label,
                type=field_type,
                options=options,
                required=required,
                tournament_id=tournament_id
            )
            db.session.add(new_field)
        db.session.commit()
        flash("Form fields added successfully.", "success")
        return redirect(url_for('tournaments.index'))

    return render_template("tournaments/add_form.html", tournaments=tournaments)

@tournaments_bp.route('/edit_form/<int:tournament_id>', methods=['GET', 'POST'])
@prevent_race_condition('edit_form', min_interval=1.5, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.index')))
def edit_form(tournament_id):
    """
    Edit existing custom form fields for a tournament (admin only).
    
    Allows admins to modify or delete existing form fields, or add new ones
    to a tournament that already has form fields.
    
    GET: Display current form fields with editing interface
    POST: Update, delete, or add form fields
    
    Access: Requires role >= 2 (Admin)
    
    Returns:
        GET: Rendered form editing interface with current fields
        POST: Redirect to tournaments index with success message
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
        
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament = Tournament.query.get_or_404(tournament_id)
    
    if request.method == 'POST':
        # Handle deletions first
        delete_ids = request.form.getlist('delete')
        for field_id in delete_ids:
            field = Form_Fields.query.get(int(field_id))
            if field and field.tournament_id == tournament_id:
                # Also delete associated responses
                Form_Responses.query.filter_by(field_id=field.id).delete()
                db.session.delete(field)
        
        # Handle updates to existing fields
        field_ids = request.form.getlist('field_id')
        labels = request.form.getlist('label')
        types = request.form.getlist('type')
        options_list = request.form.getlist('options')
        
        for i in range(len(field_ids)):
            field_id = int(field_ids[i])
            field = Form_Fields.query.get(field_id)
            if field and field.tournament_id == tournament_id:
                field.label = labels[i]
                field.type = types[i]
                field.options = options_list[i] if i < len(options_list) and options_list[i] != "" else None
                field.required = request.form.get(f'required_{i}') == '1'
        
        # Handle new fields (those without field_id)
        new_labels = request.form.getlist('new_label')
        new_types = request.form.getlist('new_type')
        new_options = request.form.getlist('new_options')
        
        for i in range(len(new_labels)):
            if new_labels[i]:  # Only create if label is provided
                new_field = Form_Fields(
                    label=new_labels[i],
                    type=new_types[i],
                    options=new_options[i] if i < len(new_options) and new_options[i] != "" else None,
                    required=request.form.get(f'new_required_{i}') == '1',
                    tournament_id=tournament_id
                )
                db.session.add(new_field)
        
        db.session.commit()
        flash("Form fields updated successfully.", "success")
        return redirect(url_for('tournaments.index'))
    
    # GET request - show existing fields
    fields = Form_Fields.query.filter_by(tournament_id=tournament_id).order_by(Form_Fields.id).all()
    return render_template("tournaments/edit_form.html", tournament=tournament, fields=fields)

@tournaments_bp.route('/signup', methods=['GET', 'POST'])
@prevent_race_condition(
    'tournament_signup',
    min_interval=1.5,
    redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.index'))
)
def signup():
    """
    Handle tournament signup for students - NOW WITH ENHANCED VALIDATION.
    
    Students select which event(s) they're competing in at a tournament,
    indicate if they're bringing a judge, fill out custom form fields,
    and optionally select a partner for partner events.
    
    IMPORTANT: Now redirects to multi-step confirmation process to ensure
    no signups are lost or submitted incorrectly.
    
    GET: Display signup form with available events and custom fields
    POST: Validate and redirect to confirmation flow (not direct submission)
    
    Features:
        - Multi-event signup (can sign up for multiple events at once)
        - Partner selection for partner events
        - Custom form responses
        - Judge request flag
        - Automatic Tournament_Judges record creation
    
    Returns:
        GET: Rendered signup form
        POST: Redirect to tournaments index with success message
    """
    tournaments = Tournament.query.all()

    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    now = datetime.now(EST)  # Get the current time in EST

    if not user_id:
        return redirect_to_login("Please log in")

    # Get all events the user is signed up for
    user_events = Event.query.join(User_Event, Event.id == User_Event.event_id).filter(User_Event.user_id == user_id).all()

    if request.method == 'POST':
        tournament_id = request.form.get('tournament_id')
        tournament = Tournament.query.get(tournament_id)
        if not tournament:
            flash("Tournament not found.", "error")
            return redirect(url_for('tournaments.signup'))

        # Ensure signup deadline hasn't passed
        if tournament.signup_deadline:
            sd = tournament.signup_deadline
            if sd.tzinfo is None:
                sd = EST.localize(sd)
            if sd < now:
                flash("The signup deadline for this tournament has passed.", "error")
                return redirect(url_for('tournaments.signup'))

        # Prevent signup if there are no form fields (no signup form)
        if not tournament.form_fields or len(tournament.form_fields) == 0:
            flash("Signup is not available for this tournament.", "error")
            return redirect(url_for('tournaments.signup'))

        print("Tournament found")

        bringing_judge = False

        # Get selected events from form
        selected_event_ids = request.form.getlist('user_event')

        # Create or update Tournament_Signups for each selected event
        for event_id in selected_event_ids:
            signup = Tournament_Signups.query.filter_by(user_id=user_id, tournament_id=tournament_id, event_id=event_id).first()
            
            # Get partner ID for this event if it's a partner event
            partner_id = request.form.get(f'partner_{event_id}')
            if partner_id:
                try:
                    partner_id = int(partner_id)
                except (ValueError, TypeError):
                    partner_id = None
            else:
                partner_id = None
            
            if not signup:
                signup = Tournament_Signups(
                    user_id=user_id,
                    tournament_id=tournament_id,
                    event_id=event_id,
                    is_going=True,
                    partner_id=partner_id,
                    created_at=datetime.now(EST)
                )
                db.session.add(signup)
            else:
                signup.is_going = True
                signup.partner_id = partner_id
                signup.created_at = datetime.now(EST)
            
            # If this is a partner event and a partner was selected, create/update the partner's signup too
            if partner_id:
                partner_signup = Tournament_Signups.query.filter_by(user_id=partner_id, tournament_id=tournament_id, event_id=event_id).first()
                if not partner_signup:
                    partner_signup = Tournament_Signups(
                        user_id=partner_id,
                        tournament_id=tournament_id,
                        event_id=event_id,
                        is_going=True,
                        partner_id=user_id,
                        created_at=datetime.now(EST)
                    )
                    db.session.add(partner_signup)
                else:
                    partner_signup.partner_id = user_id
                    if not partner_signup.is_going:
                        partner_signup.is_going = True
                        partner_signup.created_at = datetime.now(EST)

        # For each field in the selected tournament, capture the user's response
        for field in tournament.form_fields:
            field_name = f'field_{field.id}'
            response_value = request.form.get(field_name)
            # Check for the "Are you bringing a judge?" question
            if field.label.strip().lower() == "are you bringing a judge?":
                if response_value and response_value.lower() in ["yes", "true", "on", "1"]:
                    print("bringing judge")
                    bringing_judge = True
            new_response = Form_Responses(
                tournament_id=tournament.id,
                user_id=user_id,
                field_id=field.id,
                response=response_value,
                submitted_at=datetime.now(EST)
            )
            db.session.add(new_response)

        # Add Tournament_Judges rows for selected events only
        for event_id in selected_event_ids:
            # Check if Tournament_Judges entry already exists for this child/tournament/event combination
            existing_judge = Tournament_Judges.query.filter_by(
                child_id=user_id,
                tournament_id=tournament_id,
                event_id=event_id
            ).first()
            
            if not existing_judge:
                judge_acceptance = Tournament_Judges(
                    accepted=False,
                    judge_id=None,
                    child_id=user_id,
                    tournament_id=tournament_id,
                    event_id=event_id
                )
                db.session.add(judge_acceptance)

        # Commit all changes (Tournament_Signups, Form_Responses, Tournament_Judges)
        db.session.commit()

        # Handle judge selection if needed
        if bringing_judge:
            return redirect(url_for('tournaments.bringing_judge', tournament_id=tournament_id))
        
        flash("Your responses have been submitted.", "success")
        return redirect(url_for('tournaments.index'))
    else:
        # if a tournament is selected via query string, show its form fields
        tournament_id = request.args.get('tournament_id')
        selected_tournament = Tournament.query.get(tournament_id) if tournament_id else None

        # Localize signup_deadline for all tournaments and filter out expired ones
        valid_tournaments = []
        for tournament in tournaments:
            if tournament.signup_deadline:
                sd = tournament.signup_deadline
                if sd.tzinfo is None:
                    sd = EST.localize(sd)
                # attach localized deadline back to object for templates
                tournament.signup_deadline = sd
                if sd >= now:
                    valid_tournaments.append(tournament)
            else:
                # If no deadline is set, consider it valid
                valid_tournaments.append(tournament)

        fields = selected_tournament.form_fields if selected_tournament else []

        return render_template(
            "tournaments/signup.html",
            tournaments=valid_tournaments,
            selected_tournament=selected_tournament,
            fields=fields,
            now=now,
            user_events=user_events
        )

@tournaments_bp.route('/bringing_judge/<int:tournament_id>', methods=['POST', 'GET'])
@prevent_race_condition('bringing_judge', min_interval=1.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.index')))
def bringing_judge(tournament_id):
    """
    Select which judge (parent) the student is bringing to a tournament.
    
    After indicating they're bringing a judge during signup, students use
    this route to specify which of their registered parents will judge.
    Automatically updates Tournament_Signups and Tournament_Judges records.
    
    Args:
        tournament_id (int): The tournament the student is attending
    
    GET: Display form with dropdown of student's parent judges
    POST: Save judge selection and update database
    
    Judge Selection:
        - Lists all parents associated with student (from Judges table)
        - Updates Tournament_Signups.judge_id
        - Updates Tournament_Signups.bringing_judge to True
        - Updates Tournament_Judges records for all events student signed up for
    
    Database Updates:
        - Tournament_Signups: bringing_judge=True, judge_id set
        - Tournament_Judges: judge_id populated for student's events
    
    Returns:
        GET: Rendered judge selection form with parent options
        POST: Redirect to tournaments index with success message
    """
    user_id = session.get('user_id')

    if not user_id:
        flash("Log in first")
        return redirect_to_login()

    # Get all Judges entries where the current user is the child
    judges = Judges.query.filter_by(child_id=user_id).all()

    # Build a list of tuples: (judge_id, judge_name)
    judge_options = []
    for judge in judges:
        judge_user = User.query.filter_by(id=judge.judge_id).first()
        if judge_user:
            judge_options.append((judge.judge_id, f"{judge_user.first_name} {judge_user.last_name}"))

    selected_judge_id = None


    if request.method == "POST":
        selected_judge_id = request.form.get("judge_id")

        user_tournament_signup = Tournament_Signups.query.filter_by(user_id=user_id, tournament_id=tournament_id).first()

        if user_tournament_signup:
            user_tournament_signup.bringing_judge = True
            user_tournament_signup.judge_id = selected_judge_id

            # Only add Tournament_Judges rows for events the user actually signed up for in this tournament
            # Find events from Tournament_Judges where child_id=user_id and tournament_id=tournament_id and judge_id is None
            judge_rows = Tournament_Judges.query.filter_by(child_id=user_id, tournament_id=tournament_id, judge_id=None).all()
            for judge_row in judge_rows:
                judge_row.judge_id = selected_judge_id
            db.session.commit()

            flash("Judge selection saved.", "success")
            return redirect(url_for('tournaments.index'))

    return render_template(
        'tournaments/bringing_judge.html',
        judge_options=judge_options,
        selected_judge_id=selected_judge_id
    )

@tournaments_bp.route('/edit_tournament/<int:tournament_id>', methods=['GET', 'POST'])
@prevent_race_condition('edit_tournament', min_interval=2.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.index')))
def edit_tournament(tournament_id):
    """
    Edit tournament information (admin only).
    
    GET: Display tournament edit form
    POST: Update tournament with provided details
    
    Form Fields:
        - name: Tournament name
        - address: Tournament location
        - date: Tournament date and time (YYYY-MM-DDTHH:MM)
        - signup_deadline: Deadline for signups (YYYY-MM-DDTHH:MM)
        - performance_deadline: Deadline for results submission (YYYY-MM-DDTHH:MM)
    
    Access: Requires role >= 2 (Admin)
    
    Returns:
        GET: Rendered tournament edit form
        POST: Redirect to tournaments index with success message
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
    
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament = Tournament.query.get_or_404(tournament_id)
    
    if request.method == "POST":
        name = request.form.get("name")
        address = request.form.get("address")
        date_str = request.form.get("date")
        signup_deadline_str = request.form.get("signup_deadline")
        performance_deadline_str = request.form.get("performance_deadline")

        try:
            # Convert string inputs to datetime objects
            date = datetime.strptime(date_str, "%Y-%m-%dT%H:%M")
            signup_deadline = datetime.strptime(signup_deadline_str, "%Y-%m-%dT%H:%M")
            performance_deadline = datetime.strptime(performance_deadline_str, "%Y-%m-%dT%H:%M")
        except ValueError:
            flash("Invalid date format. Please use the date pickers.", "error")
            return render_template("tournaments/edit_tournament.html", tournament=tournament)

        tournament.name = name
        tournament.address = address
        tournament.date = date
        tournament.signup_deadline = signup_deadline
        tournament.performance_deadline = performance_deadline

        db.session.commit()
        flash(f"Tournament '{name}' updated successfully", "success")
        return redirect(url_for('tournaments.index'))

    return render_template("tournaments/edit_tournament.html", tournament=tournament)


@tournaments_bp.route('/delete_tournament/<int:tournament_id>', methods=['POST'])
def delete_tournament(tournament_id):
    """
    Delete a tournament and all related data (admin only).
    
    Permanently removes a tournament from the system. Related data is
    cascade-deleted through database relationships.
    
    Args:
        tournament_id (int): The ID of the tournament to delete
    
    Cascade Deletions:
        - Tournament_Signups (all student signups)
        - Tournament_Performance (all submitted results)
        - Form_Fields (custom form fields)
        - Form_Responses (student form responses)
        - Tournament_Judges (judge requests and acceptances)
    
    Access: Requires role >= 2 (Admin)
    
    Warning:
        This is a destructive operation with no confirmation dialog.
        All tournament data is permanently lost.
    
    Returns:
        Redirect to tournaments index
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
        
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))

    tournament = Tournament.query.filter_by(id=tournament_id).first()

    db.session.delete(tournament)
    db.session.commit()

    return redirect(url_for('tournaments.index'))

@tournaments_bp.route('/judge_requests', methods=['POST', 'GET'])
@prevent_race_condition('judge_requests', min_interval=1.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.index')))
def judge_requests():
    """
    View and respond to judge requests (parents only).
    
    Parents see a list of all tournaments where their children have requested
    them to judge. They can accept or decline each request.
    
    GET: Display all pending judge requests
    POST: Update acceptance status for judge requests
    
    Displayed Information:
        - Tournament name and location
        - Tournament date
        - Child's name (which child is requesting)
        - Current acceptance status
    
    Form Actions:
        - decision_{request_id}: 'yes' or 'no' for each request
        - Updates Tournament_Judges.accepted field
    
    Access: Requires user.is_parent == True
    
    Use Case:
        1. Student signs up for tournament and requests parent to judge
        2. Parent logs in and visits this route
        3. Parent reviews each request (tournament, date, child)
        4. Parent accepts or declines each request
        5. Acceptance status saved in Tournament_Judges
    
    Returns:
        GET: Rendered judge requests page with all pending requests
        POST: Redirect to judge_requests with success message
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Must Be Logged In")
        return redirect_to_login()

    if not user.is_parent:
        flash("Must be a parent")
        return redirect(url_for('main.index'))

    judge_requests = Tournament_Judges.query.filter_by(judge_id=user_id).all()

    # Prepare data for template
    judge_requests_data = []
    for req in judge_requests:
        tournament = Tournament.query.get(req.tournament_id)
        child = User.query.get(req.child_id)
        judge_requests_data.append({
            'id': req.id,
            'tournament_name': tournament.name if tournament else '',
            'address': tournament.address if tournament else '',
            'date': tournament.date if tournament else None,
            'child_name': f"{child.first_name} {child.last_name}" if child else '',
            'accepted': req.accepted,
        })

    if request.method == 'POST':
        for req in judge_requests:
            decision = request.form.get(f"decision_{req.id}")
            req.accepted = True if decision == 'yes' else False
        db.session.commit()
        flash("Decisions updated.", "success")
        return redirect(url_for('tournaments.judge_requests'))

    return render_template('tournaments/judge_requests.html', user=user, judge_requests=judge_requests_data)


@tournaments_bp.route('/my_tournaments')
def my_tournaments():
    """
    View list of tournaments the current user attended.
    
    Displays all past tournaments where the user signed up (is_going=True),
    showing submission status and allowing result submission if not already
    submitted (deadline shows as warning, not blocker).
    
    Displayed Information:
        - Tournament name, date, location
        - Performance submission status
        - Whether user can still submit results
        - Performance deadline (for reference)
    
    Submission Eligibility:
        - User must have signed up (Tournament_Signups.is_going=True)
        - User must not have already submitted results
        - Deadline passed = warning only, not blocking
    
    Features:
        - View-only access to already-submitted results
        - Direct link to submit results if not yet submitted
        - Visual indicators for submission status
        - Sorted by date (newest first)
    
    Returns:
        Rendered my_tournaments page with tournament list and submission status
    """
    user_id = session.get('user_id')
    if not user_id:
        flash("Must Be Logged In")
        return redirect_to_login()

    user = User.query.filter_by(id=user_id).first()
    tournaments = Tournament.query.order_by(Tournament.date.desc()).all()
    now = datetime.now(EST)

    # Prepare data for template: show past tournaments, allow submit if not submitted, view-only if submitted
    my_tournaments_data = []
    for tournament in tournaments:
        # Localize performance_deadline if needed
        if tournament.performance_deadline and tournament.performance_deadline.tzinfo is None:
            tournament.performance_deadline = EST.localize(tournament.performance_deadline)

        # Check if user attended (signed up and is_going)
        # Query for any signup for this tournament where is_going=True
        signup = Tournament_Signups.query.filter_by(
            user_id=user_id, 
            tournament_id=tournament.id, 
            is_going=True
        ).first()
        
        if not signup:
            continue  # Only show tournaments the user attended
        
        # Verify user has submitted form responses (matches admin view logic)
        # This ensures manufactured signups have proper form responses created
        has_form_responses = Form_Responses.query.filter_by(
            tournament_id=tournament.id,
            user_id=user_id
        ).first() is not None
        
        if not has_form_responses:
            # User has signup but no form responses - likely incomplete/broken signup
            # Skip it to avoid confusion
            continue

        # Check if user already submitted performance
        performance = Tournament_Performance.query.filter_by(user_id=user_id, tournament_id=tournament.id).first()

        # Allow submission if not already submitted (deadline is just a warning now)
        can_submit = not performance

        my_tournaments_data.append({
            'tournament': tournament,
            'performance': performance,
            'can_submit': can_submit
        })

    return render_template('tournaments/my_tournaments.html', my_tournaments=my_tournaments_data, now=now, user=user)

@tournaments_bp.route('/submit_results/<int:tournament_id>', methods=['GET', 'POST'])
@prevent_race_condition('submit_results', min_interval=2.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.view_results', tournament_id=request.view_args.get('tournament_id'))))
def submit_results(tournament_id):
    """
    Mark tournament results as finalized for roster generation (admin only).
    
    Admin route to mark a tournament's results as "finalized" by setting 
    results_submitted=True. This signals that rosters for the next tournament
    can be generated. However, students can still submit/edit results after
    this point - they just receive a warning about potential impact.
    
    Args:
        tournament_id (int): The tournament to finalize
    
    GET: Display results submission status and statistics
    POST: Mark tournament.results_submitted = True
    
    Validation:
        - Tournament date must have passed (can't finalize future tournaments)
        - Results must not already be submitted
    
    Statistics Displayed:
        - Total participants (signed up students)
        - Submitted results count
        - Pending results count
        - List of all signups
        - List of submitted performances
    
    Effect of Marking Results as Submitted:
        - Sets tournament.results_submitted = True
        - Enables roster generation for next tournament
        - Shows warning to students submitting late results
        - Does NOT prevent students from submitting/editing results
    
    Access: Requires role >= 2 (Admin)
    
    Returns:
        GET: Rendered results submission page with statistics
        POST: Redirect to view_results with success message
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
    
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:  # Only admins can submit results
        flash("You are not authorized to submit tournament results", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament = Tournament.query.get_or_404(tournament_id)
    now = datetime.now(EST)
    
    # Check if tournament date has passed
    tournament_date = tournament.date
    if tournament_date.tzinfo is None:
        tournament_date = EST.localize(tournament_date)
    
    if tournament_date >= now:
        flash("Cannot submit results for a tournament that hasn't happened yet", "error")
        return redirect(url_for('tournaments.index'))
    
    # Check if results have already been submitted
    if tournament.results_submitted:
        flash("Results have already been submitted for this tournament", "error")
        return redirect(url_for('tournaments.view_results', tournament_id=tournament_id))
    
    if request.method == 'POST':
        # Mark results as submitted (enables roster generation)
        tournament.results_submitted = True
        db.session.commit()
        
        flash("Tournament results have been finalized for roster generation. Students can still submit/edit results but will receive a warning.", "success")
        return redirect(url_for('tournaments.view_results', tournament_id=tournament_id))
    
    # Get tournament signups and performances for context
    signups = Tournament_Signups.query.filter_by(tournament_id=tournament_id, is_going=True).all()
    performances = Tournament_Performance.query.filter_by(tournament_id=tournament_id).all()
    
    # Calculate statistics
    total_participants = len(signups)
    submitted_results = len(performances)
    pending_results = total_participants - submitted_results
    
    return render_template('tournaments/submit_results.html', 
                         tournament=tournament, 
                         signups=signups,
                         performances=performances,
                         total_participants=total_participants,
                         submitted_results=submitted_results,
                         pending_results=pending_results)

@tournaments_bp.route('/view_results/<int:tournament_id>')
def view_results(tournament_id):
    """
    View all submitted results for a tournament.
    
    Displays a comprehensive list of all Tournament_Performance records
    for a specific tournament, showing which users submitted results and
    their performance details.
    
    Args:
        tournament_id (int): The tournament to view results for
    
    Displayed Information:
        - Tournament details (name, date, location)
        - List of all submitted performances with:
            * Student name
            * Points earned
            * Bid status
            * Speaker rank
            * Stage reached
    
    Access: Requires login (any authenticated user)
    
    Use Cases:
        - Admins reviewing submitted results
        - Students viewing tournament outcomes
        - Checking who has/hasn't submitted results
        - Verifying performance data before finalizing
    
    Returns:
        Rendered view_results page with all tournament performances
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
    
    tournament = Tournament.query.get_or_404(tournament_id)
    
    # Get all tournament performances for this tournament
    performances = Tournament_Performance.query.filter_by(tournament_id=tournament_id).all()
    
    # Get user details for each performance
    performance_data = []
    for performance in performances:
        user = User.query.get(performance.user_id)
        performance_data.append({
            'user': user,
            'performance': performance
        })
    
    return render_template('tournaments/view_results.html', tournament=tournament, performance_data=performance_data)

@tournaments_bp.route('/tournament_results/<int:tournament_id>', methods=['GET', 'POST'])
@prevent_race_condition('tournament_results', min_interval=2.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.view_results', tournament_id=request.view_args.get('tournament_id'))))
def tournament_results(tournament_id):
    """
    Submit tournament performance results (students).
    
    Students use this route to submit their tournament performance after
    competing. The system calculates points using a refined ranking formula.
    
    Args:
        tournament_id (int): The tournament to submit results for
    
    GET: Display result submission form (if eligible)
    POST: Process and save tournament performance
    
    Form Fields:
        - bid: 'yes' or 'no' - Did student receive a bid?
        - overall_rank: User's overall placement at tournament
        - total_competitors: Total number of competitors in the event
        - rank: Speaker rank/placement (legacy, still tracked)
        - stage: Elimination round reached (None, Doubles, Octas, Quarters, Semis, Finals)
    
    Points Calculation (New System):
        Formula: 1 + 9 * ((n - r)/(n-1))^k
        - n: total_competitors
        - r: overall_rank
        - k: decay_coefficient (default 2.0)
        
        This provides 0-10 points based on relative placement.
        
        Additional Points:
        - Bid Points:
            * First-ever bid: +15 points
            * Subsequent bids: +5 points each
        
        - Stage Points:
            * None: 0
            * Double Octafinals: 2 points
            * Octafinals: 3 points
            * Quarter Finals: 4 points
            * Semifinals: 5 points
            * Finals: 6 points
    
    Database Updates:
        - Creates Tournament_Performance record
        - Updates user.points (cumulative tournament points)
        - Updates user.bids counter if bid received
    
    Validation:
        - User must be logged in
        - User cannot have already submitted for this tournament (use edit instead)
    
    Note:
        - Tournament closure no longer blocks submissions
        - Warning shown if rosters already generated
    
    Returns:
        GET: Rendered result submission form
        POST: Redirect to user profile with updated points
    """
    # Import Tournament_Performance locally to avoid any import issues
    from mason_snd.models.tournaments import Tournament_Performance
    
    user_id = session.get('user_id')
    if not user_id:
        flash("Must Be Logged In")
        return redirect_to_login()

    tournament = Tournament.query.get_or_404(tournament_id)
    now = datetime.now(EST)
    
    # Check if rosters have been generated or deadline passed (warning, not blocking)
    roster_warning = False
    deadline_warning = False
    
    if tournament.results_submitted:
        roster_warning = True
    
    # Check if performance deadline has passed
    if tournament.performance_deadline:
        performance_deadline = tournament.performance_deadline
        if performance_deadline.tzinfo is None:
            performance_deadline = EST.localize(performance_deadline)
        if now > performance_deadline:
            deadline_warning = True
    
    # Check if user already submitted performance
    performance = Tournament_Performance.query.filter_by(user_id=user_id, tournament_id=tournament_id).first()

    if request.method == 'POST' and not performance:
        # Get form data
        bid_str = request.form.get('bid')  # 'yes' or 'no'
        stage_str = request.form.get('stage')
        overall_rank_str = request.form.get('overall_rank')
        total_competitors_str = request.form.get('total_competitors')

        # Convert bid to boolean
        bid = True if bid_str == 'yes' else False

        # Convert overall_rank and total_competitors
        try:
            overall_rank = int(overall_rank_str) if overall_rank_str else None
            total_competitors = int(total_competitors_str) if total_competitors_str else None
        except (ValueError, TypeError):
            flash("Invalid overall rank or total competitors submitted", "error")
            return redirect(request.url)

        # Validate that if one is provided, both are provided
        if (overall_rank is None) != (total_competitors is None):
            flash("Please provide both overall rank and total competitors", "error")
            return redirect(request.url)

        # Validate overall_rank is within valid range
        if overall_rank is not None and total_competitors is not None:
            if overall_rank < 1 or overall_rank > total_competitors:
                flash(f"Overall rank must be between 1 and {total_competitors}", "error")
                return redirect(request.url)

        # Convert stage to numeric value
        stage_map = {
            "None": 0,
            "Double Octafinals": 1,
            "Octafinals": 2,
            "Quarter Finals": 3,
            "Semifinals": 4,
            "Finals": 5
        }
        stage = stage_map.get(stage_str, 0)

        # Calculate points
        ranking_points = 0
        user = User.query.filter_by(id=user_id).first()
        
        # Calculate ranking points using new formula if data provided
        if overall_rank is not None and total_competitors is not None and total_competitors > 1:
            k = 2.0  # decay coefficient
            n = total_competitors
            r = overall_rank
            # Formula: 1 + 9 * ((n - r)/(n-1))^k
            ranking_points = 1 + 9 * (((n - r) / (n - 1)) ** k)
            ranking_points = round(ranking_points, 2)  # Round to 2 decimal places

        # Add bid points
        bid_points = 0
        previous_bids = Tournament_Performance.query.filter_by(user_id=user_id, bid=True).first()
        
        if bid:
            if previous_bids is None:
                # User has never received a bid before - award 15 points
                bid_points = 15
            else:
                # User has received bid(s) before - award 5 points
                bid_points = 5
        
        # Add stage points
        stage_points = 0
        if stage != 0:
            stage_points = stage + 1

        # Total points
        total_points = int(ranking_points + bid_points + stage_points)

        # Save to DB
        tournament_performance = Tournament_Performance(
            points=total_points,
            bid=bid,
            rank=None,  # Legacy field, not used in new system
            stage=stage,
            overall_rank=overall_rank,
            total_competitors=total_competitors,
            decay_coefficient=2.0 if (overall_rank is not None and total_competitors is not None) else None,
            user_id=user_id,
            tournament_id=tournament_id
        )

        user.points += total_points
        # Update user's bid count if they received a bid
        if bid:
            user.bids = (user.bids or 0) + 1
        
        db.session.add(tournament_performance)
        db.session.commit()

        flash(f"Tournament results submitted successfully! You earned {total_points} points.", "success")
        return redirect(url_for('profile.index', user_id=user_id))

    return render_template("tournaments/tournament_results.html", 
                         performance=performance, 
                         tournament=tournament,
                         roster_warning=roster_warning,
                         deadline_warning=deadline_warning)

@tournaments_bp.route('/edit_results/<int:performance_id>', methods=['GET', 'POST'])
@prevent_race_condition('edit_results', min_interval=2.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('tournaments.view_results', tournament_id=request.view_args.get('tournament_id'))))
def edit_results(performance_id):
    """
    Edit previously submitted tournament performance results.
    
    Allows users to update their tournament results. Points are recalculated
    and the difference is applied to the user's total points.
    
    Args:
        performance_id (int): The Tournament_Performance record to edit
    
    GET: Display edit form with current values
    POST: Update performance record and recalculate points
    
    Form Fields:
        - bid: 'yes' or 'no' - Did student receive a bid?
        - overall_rank: User's overall placement at tournament
        - total_competitors: Total number of competitors in the event
        - rank: Speaker rank/placement (legacy, still tracked)
        - stage: Elimination round reached
    
    Authorization:
        - Only the user who submitted can edit their own results
        - Admins cannot edit other users' results through this route
    
    Points Recalculation:
        - Old points are subtracted from user.points
        - New points are calculated using current formula
        - New points are added to user.points
        - Bid count is updated if bid status changed
    
    Returns:
        GET: Rendered edit form
        POST: Redirect to user profile with updated points
    """
    from mason_snd.models.tournaments import Tournament_Performance
    
    user_id = session.get('user_id')
    if not user_id:
        flash("Must Be Logged In", "error")
        return redirect_to_login()
    
    performance = Tournament_Performance.query.get_or_404(performance_id)
    
    # Check authorization - only the user who submitted can edit
    if performance.user_id != user_id:
        flash("You can only edit your own tournament results", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament = Tournament.query.get_or_404(performance.tournament_id)
    now = datetime.now(EST)
    
    # Check if rosters have been generated or deadline passed (warning, not blocking)
    roster_warning = tournament.results_submitted
    deadline_warning = False
    
    # Check if performance deadline has passed
    if tournament.performance_deadline:
        performance_deadline = tournament.performance_deadline
        if performance_deadline.tzinfo is None:
            performance_deadline = EST.localize(performance_deadline)
        if now > performance_deadline:
            deadline_warning = True
    
    if request.method == 'POST':
        # Get form data
        bid_str = request.form.get('bid')
        stage_str = request.form.get('stage')
        overall_rank_str = request.form.get('overall_rank')
        total_competitors_str = request.form.get('total_competitors')
        
        # Convert values
        new_bid = True if bid_str == 'yes' else False
        
        try:
            new_overall_rank = int(overall_rank_str) if overall_rank_str else None
            new_total_competitors = int(total_competitors_str) if total_competitors_str else None
        except (ValueError, TypeError):
            flash("Invalid overall rank or total competitors submitted", "error")
            return redirect(request.url)
        
        # Validate that if one is provided, both are provided
        if (new_overall_rank is None) != (new_total_competitors is None):
            flash("Please provide both overall rank and total competitors", "error")
            return redirect(request.url)
        
        # Validate overall_rank is within valid range
        if new_overall_rank is not None and new_total_competitors is not None:
            if new_overall_rank < 1 or new_overall_rank > new_total_competitors:
                flash(f"Overall rank must be between 1 and {new_total_competitors}", "error")
                return redirect(request.url)
        
        # Convert stage
        stage_map = {
            "None": 0,
            "Double Octafinals": 1,
            "Octafinals": 2,
            "Quarter Finals": 3,
            "Semifinals": 4,
            "Finals": 5
        }
        new_stage = stage_map.get(stage_str, 0)
        
        # Get user
        user = User.query.filter_by(id=user_id).first()
        
        # Store old values for point adjustment
        old_points = performance.points or 0
        old_bid = performance.bid
        
        # Calculate new ranking points using new formula if data provided
        ranking_points = 0
        if new_overall_rank is not None and new_total_competitors is not None and new_total_competitors > 1:
            k = 2.0  # decay coefficient
            n = new_total_competitors
            r = new_overall_rank
            ranking_points = 1 + 9 * (((n - r) / (n - 1)) ** k)
            ranking_points = round(ranking_points, 2)
        
        # Calculate bid points
        bid_points = 0
        # Check for previous bids excluding this performance
        previous_bids = Tournament_Performance.query.filter(
            Tournament_Performance.user_id == user_id,
            Tournament_Performance.bid == True,
            Tournament_Performance.id != performance_id
        ).first()
        
        if new_bid:
            if previous_bids is None:
                bid_points = 15  # First bid
            else:
                bid_points = 5  # Subsequent bid
        
        # Calculate stage points
        stage_points = 0
        if new_stage != 0:
            stage_points = new_stage + 1
        
        # Total new points
        new_total_points = int(ranking_points + bid_points + stage_points)
        
        # Update user points (remove old, add new)
        user.points = (user.points or 0) - old_points + new_total_points
        
        # Update bid count if bid status changed
        if old_bid != new_bid:
            if new_bid:
                user.bids = (user.bids or 0) + 1
            else:
                user.bids = max(0, (user.bids or 0) - 1)
        
        # Update performance record
        performance.bid = new_bid
        performance.rank = None  # Legacy field, not used in new system
        performance.stage = new_stage
        performance.overall_rank = new_overall_rank
        performance.total_competitors = new_total_competitors
        performance.decay_coefficient = 2.0 if (new_overall_rank is not None and new_total_competitors is not None) else None
        performance.points = new_total_points
        
        db.session.commit()
        
        flash(f"Tournament results updated successfully! Your points changed by {new_total_points - old_points:+d}.", "success")
        return redirect(url_for('profile.index', user_id=user_id))
    
    return render_template("tournaments/edit_results.html",
                         performance=performance,
                         tournament=tournament,
                         roster_warning=roster_warning,
                         deadline_warning=deadline_warning)

@tournaments_bp.route('/delete_results/<int:performance_id>', methods=['POST'])
@prevent_race_condition('delete_results', min_interval=2.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('profile.index', user_id=uid)))
def delete_results(performance_id):
    """
    Delete previously submitted tournament performance results.
    
    Allows users to remove their tournament results. Points are subtracted
    from user's total and bid count is decremented if applicable.
    
    Args:
        performance_id (int): The Tournament_Performance record to delete
    
    POST: Delete performance record and adjust user points
    
    Authorization:
        - Only the user who submitted can delete their own results
    
    Points Adjustment:
        - Subtracts performance.points from user.points
        - Decrements user.bids if performance had a bid
    
    Returns:
        POST: Redirect to user profile
    """
    from mason_snd.models.tournaments import Tournament_Performance
    
    user_id = session.get('user_id')
    if not user_id:
        flash("Must Be Logged In", "error")
        return redirect_to_login()
    
    performance = Tournament_Performance.query.get_or_404(performance_id)
    
    # Check authorization - only the user who submitted can delete
    if performance.user_id != user_id:
        flash("You can only delete your own tournament results", "error")
        return redirect(url_for('tournaments.index'))
    
    # Get user
    user = User.query.filter_by(id=user_id).first()
    
    # Adjust user points
    old_points = performance.points or 0
    user.points = (user.points or 0) - old_points
    
    # Adjust bid count if applicable
    if performance.bid:
        user.bids = max(0, (user.bids or 0) - 1)
    
    # Delete the performance
    db.session.delete(performance)
    db.session.commit()
    
    flash(f"Tournament results deleted successfully. {old_points} points removed.", "success")
    return redirect(url_for('profile.index', user_id=user_id))

@tournaments_bp.route('/search_partners')
def search_partners():
    """
    AJAX endpoint to search for tournament partners.
    
    Provides autocomplete/search functionality for finding partners for
    partner events (e.g., Public Forum debate). Returns JSON list of users
    who match the search query and are enrolled in the specified event.
    
    Query Parameters:
        - q: Search query (first name, last name, or full name)
        - event_id: Optional event filter (only show users in this event)
    
    Search Logic:
        - Minimum 2 characters required
        - Case-insensitive partial matching
        - Searches first name, last name, and full name combinations
        - Excludes current user from results
        - Limits to 10 results
        - Optionally filters to users enrolled in specific event
    
    Response JSON:
        {
            'users': [
                {
                    'id': user_id,
                    'first_name': 'John',
                    'last_name': 'Doe'
                },
                ...
            ]
        }
    
    HTTP Status Codes:
        - 200: Success with user list
        - 401: Not authenticated
    
    Use Case:
        - Frontend AJAX call for partner autocomplete
        - Real-time search as user types
        - Ensures partner is eligible (enrolled in same event)
    
    Returns:
        JSON response with matching users array
    """
    from flask import jsonify
    
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'error': 'Not logged in'}), 401
    
    query = request.args.get('q', '').strip()
    event_id = request.args.get('event_id')
    
    if not query or len(query) < 2:
        return jsonify({'users': []})
    
    # Search for users by name, excluding current user
    users = User.query.filter(
        db.or_(
            User.first_name.ilike(f'%{query}%'),
            User.last_name.ilike(f'%{query}%'),
            db.func.concat(User.first_name, ' ', User.last_name).ilike(f'%{query}%')
        ),
        User.id != user_id  # Exclude current user
    ).limit(10).all()
    
    # Filter users who are signed up for the same event
    if event_id:
        from mason_snd.models.events import User_Event
        event_users = User_Event.query.filter_by(event_id=event_id, active=True).all()
        event_user_ids = [eu.user_id for eu in event_users]
        users = [user for user in users if user.id in event_user_ids]
    
    return jsonify({
        'users': [
            {
                'id': user.id,
                'first_name': user.first_name,
                'last_name': user.last_name
            }
            for user in users
        ]
    })

@tournaments_bp.route('/view_form_responses/<int:tournament_id>')
def view_form_responses(tournament_id):
    user_id = session.get('user_id')
    
    if not user_id:
        return redirect_to_login("Please log in")
    
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament = Tournament.query.get_or_404(tournament_id)
    
    form_fields = Form_Fields.query.filter_by(tournament_id=tournament_id).order_by(Form_Fields.id).all()
    
    if not form_fields:
        flash(f"No form fields found for {tournament.name}", "warning")
        return redirect(url_for('tournaments.index'))
    
    # Only show signups where is_going=True (confirmed attendees)
    signups = Tournament_Signups.query.filter_by(tournament_id=tournament_id, is_going=True).all()
    
    user_responses = {}
    user_responses_json = {}
    
    for signup in signups:
        user_obj = User.query.get(signup.user_id)
        if not user_obj:
            continue
            
        if signup.user_id not in user_responses:
            user_responses[signup.user_id] = {
                'user': user_obj,
                'signup': signup,
                'responses': {}
            }
            user_responses_json[signup.user_id] = {
                'user': {
                    'first_name': user_obj.first_name,
                    'last_name': user_obj.last_name,
                    'email': user_obj.email
                },
                'signup': {
                    'created_at': signup.created_at.strftime('%Y-%m-%d %H:%M:%S') if signup.created_at else ''
                },
                'responses': {}
            }
        
        responses = Form_Responses.query.filter_by(
            tournament_id=tournament_id,
            user_id=signup.user_id
        ).all()
        
        for response in responses:
            field = Form_Fields.query.get(response.field_id)
            if field:
                user_responses[signup.user_id]['responses'][field.id] = response.response
                user_responses_json[signup.user_id]['responses'][field.id] = response.response
    
    form_fields_json = [
        {
            'id': field.id,
            'label': field.label,
            'type': field.type,
            'required': field.required
        }
        for field in form_fields
    ]
    
    return render_template('tournaments/view_form_responses.html',
                         tournament=tournament,
                         form_fields=form_fields,
                         form_fields_json=form_fields_json,
                         user_responses=user_responses,
                         user_responses_json=user_responses_json,
                         user=user)


@tournaments_bp.route('/download_form_responses/<int:tournament_id>')
def download_form_responses(tournament_id):
    user_id = session.get('user_id')
    
    if not user_id:
        return redirect_to_login("Please log in")
    
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))
    
    if pd is None or openpyxl is None:
        flash("Excel functionality not available. Please install pandas and openpyxl.", "error")
        return redirect(url_for('tournaments.view_form_responses', tournament_id=tournament_id))
    
    tournament = Tournament.query.get_or_404(tournament_id)
    
    form_fields = Form_Fields.query.filter_by(tournament_id=tournament_id).order_by(Form_Fields.id).all()
    
    if not form_fields:
        flash(f"No form fields found for {tournament.name}", "warning")
        return redirect(url_for('tournaments.view_form_responses', tournament_id=tournament_id))
    
    # Only include signups where is_going=True (confirmed attendees)
    signups = Tournament_Signups.query.filter_by(tournament_id=tournament_id, is_going=True).all()
    
    response_data = []
    
    for signup in signups:
        user_obj = User.query.get(signup.user_id)
        if not user_obj:
            continue
        
        row = {
            'Signup Timestamp': signup.created_at.strftime('%Y-%m-%d %H:%M:%S') if signup.created_at else '',
            'Student Name': f"{user_obj.first_name} {user_obj.last_name}",
            'Email': user_obj.email
        }
        
        responses = Form_Responses.query.filter_by(
            tournament_id=tournament_id,
            user_id=signup.user_id
        ).all()
        
        response_dict = {r.field_id: r.response for r in responses}
        
        for field in form_fields:
            row[field.label] = response_dict.get(field.id, '')
        
        response_data.append(row)
    
    if not response_data:
        flash(f"No form responses found for {tournament.name}", "warning")
        return redirect(url_for('tournaments.view_form_responses', tournament_id=tournament_id))
    
    df = pd.DataFrame(response_data)
    
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    df.to_excel(writer, sheet_name=f'{tournament.name} Responses', index=False)
    
    from openpyxl.styles import PatternFill, Font, Alignment
    
    workbook = writer.book
    worksheet = workbook.active
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    writer.close()
    output.seek(0)
    
    safe_tournament_name = "".join(c for c in tournament.name if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_tournament_name = safe_tournament_name.replace(' ', '_')
    filename = f"{safe_tournament_name}_form_responses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(output, 
                     as_attachment=True, 
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@tournaments_bp.route('/download_ranked_signups/<int:tournament_id>')
def download_ranked_signups(tournament_id):
    user_id = session.get('user_id')
    
    if not user_id:
        return redirect_to_login("Please log in")
    
    user = User.query.filter_by(id=user_id).first()
    if not user or user.role < 2:
        flash("You are not authorized to access this page", "error")
        return redirect(url_for('tournaments.index'))
    
    if pd is None or openpyxl is None:
        flash("Excel functionality not available. Please install pandas and openpyxl.", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament = Tournament.query.get_or_404(tournament_id)
    
    signups = Tournament_Signups.query.filter_by(tournament_id=tournament_id, is_going=True).all()
    
    if not signups:
        flash(f"No signups found for {tournament.name}", "warning")
        return redirect(url_for('tournaments.index'))
    
    from mason_snd.models.metrics import MetricsSettings
    
    settings = MetricsSettings.query.first()
    if not settings:
        settings = MetricsSettings()
        db.session.add(settings)
        db.session.commit()
    tournament_weight = settings.tournament_weight
    effort_weight = settings.effort_weight
    
    ranked_data = []
    
    for signup in signups:
        user_obj = User.query.get(signup.user_id)
        if not user_obj:
            continue
        
        event = Event.query.get(signup.event_id) if signup.event_id else None
        event_name = event.event_name if event else 'Unknown Event'
        
        event_type = 'Unknown'
        if event:
            if event.event_type == 0:
                event_type = 'Speech'
            elif event.event_type == 1:
                event_type = 'LD'
            elif event.event_type == 2:
                event_type = 'PF'
        
        partner = User.query.get(signup.partner_id) if signup.partner_id else None
        partner_name = f"{partner.first_name} {partner.last_name}" if partner else ''
        
        tournament_pts = user_obj.tournament_points if hasattr(user_obj, 'tournament_points') else 0
        effort_pts = user_obj.effort_points if hasattr(user_obj, 'effort_points') else 0
        weighted_pts = round((tournament_pts * tournament_weight) + (effort_pts * effort_weight), 2)
        
        ranked_data.append({
            'Event': event_name,
            'Category': event_type,
            'Competitor Name': f"{user_obj.first_name} {user_obj.last_name}",
            'Partner': partner_name,
            'Weighted Points': weighted_pts,
            'Tournament Points': tournament_pts,
            'Effort Points': effort_pts,
            'Email': user_obj.email,
            'User ID': user_obj.id,
            'Event ID': signup.event_id if signup.event_id else ''
        })
    
    ranked_data.sort(key=lambda x: (-x['Weighted Points'], x['Event'], x['Competitor Name']))
    
    event_groups = {}
    for row in ranked_data:
        event_key = (row['Event'], row['Event ID'])
        if event_key not in event_groups:
            event_groups[event_key] = []
        event_groups[event_key].append(row)
    
    for event_key, rows in event_groups.items():
        rows.sort(key=lambda x: -x['Weighted Points'])
        for idx, row in enumerate(rows, 1):
            row['Rank'] = idx
    
    final_data = []
    for event_key in sorted(event_groups.keys(), key=lambda x: x[0]):
        final_data.extend(event_groups[event_key])
    
    column_order = [
        'Rank', 'Event', 'Category', 'Competitor Name', 'Partner', 
        'Weighted Points', 'Tournament Points', 'Effort Points', 
        'Email', 'User ID', 'Event ID'
    ]
    
    df = pd.DataFrame(final_data, columns=column_order)
    
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    df.to_excel(writer, sheet_name='Ranked Signups', index=False)
    
    from openpyxl.styles import PatternFill, Font, Alignment
    
    workbook = writer.book
    worksheet = workbook.active
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    writer.close()
    output.seek(0)
    
    safe_tournament_name = "".join(c for c in tournament.name if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_tournament_name = safe_tournament_name.replace(' ', '_')
    filename = f"{safe_tournament_name}_ranked_signups_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(output, 
                     as_attachment=True, 
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')




@tournaments_bp.route('/signup/requirements/<int:tournament_id>')
def signup_requirements(tournament_id):
    """Pre-signup requirements check screen.
    
    Shows user a checklist of requirements they must meet before
    they can proceed to the signup form.
    """
    from mason_snd.utils.tournament_signup_validator import get_signup_requirements_summary
    
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in to sign up for tournaments")
    
    requirements = get_signup_requirements_summary(user_id, tournament_id)
    
    if 'error' in requirements:
        flash(requirements['error'], 'error')
        return redirect(url_for('tournaments.index'))
    
    return render_template(
        'tournaments/signup_requirements.html',
        requirements=requirements,
        tournament_id=tournament_id
    )


@tournaments_bp.route('/signup/confirm', methods=['POST'])
def signup_confirmation():
    """First confirmation screen - review all info and initial confirmations."""
    from mason_snd.utils.tournament_signup_validator import TournamentSignupValidator
    import json
    
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
    
    tournament_id = request.form.get('tournament_id')
    if not tournament_id:
        flash("No tournament selected", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament = Tournament.query.get(tournament_id)
    if not tournament:
        flash("Tournament not found", "error")
        return redirect(url_for('tournaments.index'))
    
    # Gather all form data
    selected_event_ids = [int(eid) for eid in request.form.getlist('user_event')]
    
    # Build partners dict
    partners = {}
    for event_id in selected_event_ids:
        partner_id = request.form.get(f'partner_{event_id}')
        if partner_id:
            try:
                partners[event_id] = int(partner_id)
            except (ValueError, TypeError):
                pass
    
    # Build form responses dict
    form_responses = {}
    for field in tournament.form_fields:
        response = request.form.get(f'field_{field.id}')
        # Store the response even if None - validator will handle it
        # Convert None to empty string to avoid issues
        form_responses[field.id] = response if response is not None else ''
    
    # DEBUG: Print detailed form data
    print("\n" + "="*80)
    print("TOURNAMENT SIGNUP DEBUG INFO")
    print("="*80)
    print(f"Tournament ID: {tournament_id}")
    print(f"Tournament Name: {tournament.name}")
    print(f"User ID: {user_id}")
    print(f"\nSelected Event IDs: {selected_event_ids}")
    print(f"\nPartners: {partners}")
    print(f"\nForm Fields and Responses:")
    for field in tournament.form_fields:
        response = form_responses.get(field.id, 'NOT FOUND')
        print(f"  Field ID {field.id} - '{field.label}' (required={field.required}, type={field.type})")
        print(f"    Response: '{response}' (length: {len(str(response)) if response else 0})")
    print("="*80 + "\n")
    
    # Validate the signup data
    validator = TournamentSignupValidator(user_id, tournament_id)
    validation_result = validator.validate_signup_request({
        'selected_event_ids': selected_event_ids,
        'form_responses': form_responses,
        'partners': partners
    })
    
    # DEBUG: Print validation results
    print("\n" + "="*80)
    print("VALIDATION RESULTS")
    print("="*80)
    print(f"Is Valid: {validation_result.is_valid}")
    print(f"\nErrors ({len(validation_result.errors)}):")
    for error in validation_result.errors:
        print(f"  - Field: {error.field}")
        print(f"    Message: {error.message}")
        print(f"    Fix: {error.fix_instructions}")
    print(f"\nWarnings ({len(validation_result.warnings)}):")
    for warning in validation_result.warnings:
        print(f"  - {warning.message}")
    print(f"\nRequirements Met:")
    for req, met in validation_result.requirements_met.items():
        print(f"  {req}: {met}")
    print("="*80 + "\n")
    
    if not validation_result.is_valid:
        return render_template(
            'tournaments/signup_error.html',
            validation_result=validation_result,
            tournament_id=tournament_id
        )
    
    # Build event info for display
    events_info = []
    for event_id in selected_event_ids:
        event = Event.query.get(event_id)
        if event:
            event_info = {
                'id': event_id,
                'name': event.event_name,
                'emoji': event.event_emoji,
                'partner': None
            }
            
            if event_id in partners:
                partner = User.query.get(partners[event_id])
                if partner:
                    event_info['partner'] = {
                        'id': partner.id,
                        'name': f"{partner.first_name} {partner.last_name}",
                        'email': partner.email
                    }
            
            events_info.append(event_info)
    
    # Build signup data structure - keep integer keys for template
    signup_data = {
        'tournament_id': tournament_id,
        'events': events_info,
        'form_responses': form_responses
    }
    
    # Build JSON-serializable version with string keys
    signup_data_for_json = {
        'tournament_id': tournament_id,
        'events': events_info,
        'form_responses': {str(k): v for k, v in form_responses.items()}
    }
    
    signup_data_json = json.dumps(signup_data_for_json)
    form_fields = {field.id: field for field in tournament.form_fields}
    
    return render_template(
        'tournaments/signup_confirmation.html',
        tournament=tournament,
        signup_data=signup_data,
        signup_data_json=signup_data_json,
        form_fields=form_fields,
        validation_result=validation_result
    )


@tournaments_bp.route('/signup/final_confirm', methods=['POST'])
def signup_final_confirm():
    """Second confirmation screen - final warnings before submit."""
    import json
    
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
    
    signup_data_json = request.form.get('signup_data_json')
    if not signup_data_json:
        flash("Session expired. Please start signup again.", "error")
        return redirect(url_for('tournaments.index'))
    
    try:
        signup_data = json.loads(signup_data_json)
    except json.JSONDecodeError:
        flash("Invalid signup data. Please start again.", "error")
        return redirect(url_for('tournaments.index'))
    
    tournament_id = signup_data.get('tournament_id')
    tournament = Tournament.query.get(tournament_id)
    
    if not tournament:
        flash("Tournament not found", "error")
        return redirect(url_for('tournaments.index'))
    
    # Verify confirmations from first screen
    required_confirmations = ['confirm_info_accurate', 'confirm_commitment']
    for conf in required_confirmations:
        if not request.form.get(conf):
            flash("You must check all confirmation boxes to proceed", "error")
            return redirect(url_for('tournaments.signup', tournament_id=tournament_id))
    
    return render_template(
        'tournaments/signup_final_confirmation.html',
        tournament=tournament,
        signup_data=signup_data,
        signup_data_json=signup_data_json
    )


@tournaments_bp.route('/signup/submit', methods=['POST'])
def signup_submit():
    """Actually process and save the signup to database with transaction safety."""
    import json
    import hashlib
    
    user_id = session.get('user_id')
    if not user_id:
        return redirect_to_login("Please log in")
    
    signup_data_json = request.form.get('signup_data_json')
    if not signup_data_json:
        flash("Session expired. Please start signup again.", "error")
        return redirect(url_for('tournaments.index'))
    
    try:
        signup_data = json.loads(signup_data_json)
    except json.JSONDecodeError:
        flash("Invalid signup data. Please start again.", "error")
        return redirect(url_for('tournaments.index'))
    
    # Verify all final confirmations
    required_final_confirmations = [
        'final_confirm_reviewed',
        'final_confirm_no_mistakes',
        'final_confirm_understand_consequences'
    ]
    for conf in required_final_confirmations:
        if not request.form.get(conf):
            flash("You must check all confirmation boxes to submit", "error")
            tournament_id = signup_data.get('tournament_id')
            return redirect(url_for('tournaments.signup', tournament_id=tournament_id))
    
    tournament_id = signup_data.get('tournament_id')
    tournament = Tournament.query.get(tournament_id)
    
    if not tournament:
        flash("Tournament not found", "error")
        return redirect(url_for('tournaments.index'))
    
    # Re-validate one more time
    from mason_snd.utils.tournament_signup_validator import TournamentSignupValidator
    
    selected_event_ids = [event['id'] for event in signup_data.get('events', [])]
    partners = {}
    for event in signup_data.get('events', []):
        if event.get('partner'):
            partners[event['id']] = event['partner']['id']
    
    # Convert form_responses keys back to integers (they become strings in JSON)
    form_responses_raw = signup_data.get('form_responses', {})
    form_responses = {}
    for key, value in form_responses_raw.items():
        try:
            form_responses[int(key)] = value
        except (ValueError, TypeError):
            form_responses[key] = value
    
    # DEBUG: Print form responses conversion
    print("\n" + "="*80)
    print("FORM RESPONSES CONVERSION DEBUG")
    print("="*80)
    print(f"Raw form_responses from JSON: {form_responses_raw}")
    print(f"Converted form_responses: {form_responses}")
    print("="*80 + "\n")
    
    validator = TournamentSignupValidator(user_id, tournament_id)
    validation_result = validator.validate_signup_request({
        'selected_event_ids': selected_event_ids,
        'form_responses': form_responses,
        'partners': partners
    })
    
    if not validation_result.is_valid:
        flash("Validation failed. Please try again.", "error")
        return render_template(
            'tournaments/signup_error.html',
            validation_result=validation_result,
            tournament_id=tournament_id
        )
    
    # Begin database transaction
    now = datetime.now(EST)
    created_signups = []
    
    try:
        # Create Tournament_Signups for each event
        for event_data in signup_data.get('events', []):
            event_id = event_data['id']
            partner_id = event_data.get('partner', {}).get('id') if event_data.get('partner') else None
            
            signup = Tournament_Signups.query.filter_by(
                user_id=user_id,
                tournament_id=tournament_id,
                event_id=event_id
            ).first()
            
            if not signup:
                signup = Tournament_Signups(
                    user_id=user_id,
                    tournament_id=tournament_id,
                    event_id=event_id,
                    is_going=True,
                    partner_id=partner_id,
                    created_at=now
                )
                db.session.add(signup)
            else:
                signup.is_going = True
                signup.partner_id = partner_id
                signup.created_at = now
            
            db.session.flush()
            created_signups.append({
                'event_id': event_id,
                'event_name': event_data['name'],
                'event_emoji': event_data.get('emoji', ''),
                'signup_id': signup.id,
                'partner': event_data.get('partner')
            })
            
            # If partner event, create/update partner's signup
            if partner_id:
                partner_signup = Tournament_Signups.query.filter_by(
                    user_id=partner_id,
                    tournament_id=tournament_id,
                    event_id=event_id
                ).first()
                
                if not partner_signup:
                    partner_signup = Tournament_Signups(
                        user_id=partner_id,
                        tournament_id=tournament_id,
                        event_id=event_id,
                        is_going=True,
                        partner_id=user_id,
                        created_at=now
                    )
                    db.session.add(partner_signup)
                else:
                    partner_signup.partner_id = user_id
                    if not partner_signup.is_going:
                        partner_signup.is_going = True
                        partner_signup.created_at = now
            
            # Create Tournament_Judges entry
            existing_judge = Tournament_Judges.query.filter_by(
                child_id=user_id,
                tournament_id=tournament_id,
                event_id=event_id
            ).first()
            
            if not existing_judge:
                judge_entry = Tournament_Judges(
                    accepted=False,
                    judge_id=None,
                    child_id=user_id,
                    tournament_id=tournament_id,
                    event_id=event_id
                )
                db.session.add(judge_entry)
        
        # Create Form_Responses
        for field_id, response in form_responses.items():
            old_response = Form_Responses.query.filter_by(
                tournament_id=tournament_id,
                user_id=user_id,
                field_id=field_id
            ).first()
            
            if old_response:
                db.session.delete(old_response)
            
            new_response = Form_Responses(
                tournament_id=tournament_id,
                user_id=user_id,
                field_id=field_id,
                response=response,
                submitted_at=now
            )
            db.session.add(new_response)
        
        # Commit all changes atomically
        db.session.commit()
        
        # Generate confirmation ID
        confirmation_string = f"{user_id}-{tournament_id}-{now.isoformat()}"
        confirmation_id = hashlib.sha256(confirmation_string.encode()).hexdigest()[:16].upper()
        transaction_id = hashlib.sha256(f"{confirmation_string}-{len(created_signups)}".encode()).hexdigest()[:24].upper()
        
        return render_template(
            'tournaments/signup_success.html',
            tournament=tournament,
            events_registered=created_signups,
            confirmation_id=confirmation_id,
            transaction_id=transaction_id,
            submitted_at=now,
            signup_count=len(created_signups)
        )
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR during signup submission: {str(e)}")
        flash(
            f"An error occurred while saving your signup: {str(e)}. "
            "Please try again. If the problem persists, contact an administrator.",
            "error"
        )
        return redirect(url_for('tournaments.signup', tournament_id=tournament_id))


# Enhanced signup routes are now registered above directly in this file
