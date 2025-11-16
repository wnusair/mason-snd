"""Rosters Blueprint - Intelligent tournament roster generation and management.

Automated roster generation system that fairly distributes competitors based on judge commitments,
weighted points rankings, and event types. Includes drop penalties, partner tracking, Excel
export/import with smart name reconciliation, and publishing system for user notifications.

Roster Generation Algorithm:
    Judge Commitments:
        - Each judge brings a fixed number of competitors based on event type:
          * Speech (type 0): 6 competitors
          * Lincoln-Douglas/LD (type 1): 2 competitors
          * Public Forum/PF (type 2): 4 competitors
    
    Competitor Selection:
        1. Rank all signups within each event by weighted_points (from metrics system)
        2. Apply drop penalties (users with drops > 0 are skipped, drops decremented)
        3. Event-specific selection algorithms:
           
           Speech Events (Rotation with Random):
               - Rotate through speech events selecting top-ranked competitors
               - Every 5th selection (if 4+ speech judges): pick from middle third (randomness)
               - Tracks random selections for transparency
           
           LD/PF Events (Top-N or Middle):
               - If 1 judge: select top N competitors by rank
               - If 2+ judges: select from middle of rankings (fairness)
               - N = 2 for LD, N = 4 for PF per judge
    
    Drop Penalty System:
        - Users with drops > 0 are excluded from roster selection
        - Each exclusion decrements user's drop count by 1
        - Penalty entries tracked in Roster_Penalty_Entries
        - Replacement user tracked (next user without drops)

Key Features:
    - Automated roster generation from tournament signups and judges
    - Drop penalty enforcement with tracking
    - Partner event support (PF partner matching)
    - Excel export with multiple sheets (judges, rank view, per-event views)
    - Excel import with smart name reconciliation (ID priority, fuzzy matching)
    - Publishing system (marks rosters visible to users, creates notifications)
    - Weighted points integration (uses metrics system configuration)

Route Organization:
    Generation & Viewing:
        - index(): List upcoming tournaments and saved rosters
        - view_tournament(tournament_id): Generate and preview roster for tournament
        - save_tournament(tournament_id): Save generated roster to database
        - view_roster(roster_id): View saved roster
    
    Publishing:
        - publish_roster(roster_id): Make roster visible to users (creates notifications)
        - unpublish_roster(roster_id): Hide roster from users
    
    Excel Operations:
        - download_tournament(tournament_id): Export live tournament roster to Excel
        - download_roster(roster_id): Export saved roster to Excel (fully editable)
        - upload_roster(): Import Excel file to create/update roster (smart reconciliation)
    
    Management:
        - rename_roster(roster_id): Change roster name
        - delete_roster(roster_id): Delete roster and associated data

Excel Format:
    Sheets:
        1. Judges: Judge Name, Child, Event, Category, Number People Bringing, IDs
        2. Rank View: Rank, Competitor, Partner, Weighted Points, Event, Category, IDs
        3. Event Sheets: Per-event competitor lists with rankings
    
    Smart Reconciliation (Import):
        - ID Priority: User ID and Event ID take precedence (name changes ignored)
        - Fuzzy Name Matching: Case-insensitive name matching if ID missing
        - Change Tracking: Logs added competitors, judges, and warnings

Dependencies:
    - pandas & openpyxl: Excel functionality (optional, graceful fallback)
    - Metrics system: Weighted points calculation
    - Tournament Judges: Judge commitments and event assignments
    - Tournament Signups: Competitor availability (is_going=True)
"""

import csv
from io import StringIO, BytesIO
from math import ceil
import random

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, Response, send_file

from mason_snd.extensions import db
from mason_snd.models.auth import User, User_Published_Rosters, Roster_Penalty_Entries
from mason_snd.models.admin import User_Requirements, Requirements
from mason_snd.models.tournaments import Tournament, Tournament_Performance, Tournament_Judges, Tournament_Signups
from mason_snd.models.events import Event, User_Event, Effort_Score
from mason_snd.models.metrics import MetricsSettings
from mason_snd.blueprints.metrics.metrics import get_point_weights
from mason_snd.utils.race_protection import prevent_race_condition
from mason_snd.utils.auth_helpers import redirect_to_login

# Create a new Roster entry
from mason_snd.models.rosters import Roster, Roster_Competitors, Roster_Judge
from mason_snd.models.tournaments import Tournament_Judges
from datetime import datetime
from sqlalchemy import asc, desc, func

from datetime import datetime
import pytz


try:
    import pandas as pd
    import openpyxl
except ImportError:
    pd = None
    openpyxl = None

# Timezone constant used throughout the app
EST = pytz.timezone('US/Eastern')


rosters_bp = Blueprint('rosters', __name__, template_folder='templates')

def calculate_weighted_points(user):
    """Calculate weighted points for a user, including drop penalties.
    
    Now uses the User.weighted_points property which automatically applies
    drop penalties (-10 points per drop).
    
    Args:
        user: User object or None
    
    Returns:
        float: Weighted points with drop penalties applied, or 0 if no user
    """
    if not user:
        return 0
    return getattr(user, 'weighted_points', 0)

@rosters_bp.route('/')
def index():
    """Rosters dashboard showing upcoming tournaments and saved rosters.
    
    Displays two lists:
    - Upcoming tournaments (signup_deadline > current time) for roster generation
    - All saved rosters for viewing/management
    
    Access Control:
        Open to all users (viewing only). Generation requires admin access.
    
    Template:
        rosters/index.html with upcoming_tournaments and rosters lists.
    """
    tournaments = Tournament.query.all()
    rosters = Roster.query.all()

    upcoming_tournaments = []
    now = datetime.now(EST)
    for tournament in tournaments:
        # Keep tournaments available for roster actions up until the tournament date
        tournament_date = tournament.date
        if tournament_date is None:
            continue
        # Localize naive datetimes to EST for correct comparison
        if tournament_date.tzinfo is None:
            tournament_date = EST.localize(tournament_date)
        if tournament_date >= now:
            upcoming_tournaments.append(tournament)

    return render_template('rosters/index.html', upcoming_tournaments=upcoming_tournaments, tournaments=tournaments, rosters=rosters)



# BIG VIEW TOURNAMENT BLOCK!!!!

def get_roster_count(tournament_id):
    """Calculate total competitor spots based on judge commitments.
    
    Each judge brings a fixed number of competitors based on their event type:
    - Speech (event_type=0): 6 competitors per judge
    - Lincoln-Douglas (event_type=1): 2 competitors per judge
    - Public Forum (event_type=2): 4 competitors per judge
    
    Args:
        tournament_id (int): Tournament primary key.
    
    Returns:
        tuple: (speech_competitors, LD_competitors, PF_competitors, spots_per_event) - total spots available.
            spots_per_event: dict mapping event_id to number of spots for that specific event
    
    Note:
        Only counts accepted judges (Tournament_Judges.accepted=True).
    """
    judges = Tournament_Judges.query.filter_by(tournament_id=tournament_id, accepted=True).all()

    speech_competitors = 0
    LD_competitors = 0
    PF_competitors = 0
    spots_per_event = {}

    for judge in judges:
        event_id = judge.event_id

        event = Event.query.filter_by(id=event_id).first()

        if event.event_type == 0:
            speech_competitors += 6
            spots_per_event[event_id] = spots_per_event.get(event_id, 0) + 6
        elif event.event_type == 1:
            LD_competitors += 2
            spots_per_event[event_id] = spots_per_event.get(event_id, 0) + 2
        else:
            PF_competitors += 4
            spots_per_event[event_id] = spots_per_event.get(event_id, 0) + 4

    return speech_competitors, LD_competitors, PF_competitors, spots_per_event

# Helper: Get all signups for a tournament, grouped by event
from mason_snd.models.tournaments import Tournament_Signups

def get_signups_by_event(tournament_id):
    """Group tournament signups by event.
    
    Retrieves all confirmed signups (is_going=True) where users are actively
    enrolled in the event (User_Event.active=True) and organizes them
    into a dictionary keyed by event_id.
    
    Args:
        tournament_id (int): Tournament primary key.
    
    Returns:
        dict: {event_id: [Tournament_Signups, ...]} - signups grouped by event.
    
    Note:
        Only includes signups with:
        - is_going=True (confirmed attendance)
        - User_Event.active=True (active enrollment in the event)
    
    Bug Fix:
        Previously didn't check User_Event.active, causing users to appear
        in roster for events they were no longer active in.
    """
    signups = Tournament_Signups.query.join(
        User_Event,
        db.and_(
            User_Event.user_id == Tournament_Signups.user_id,
            User_Event.event_id == Tournament_Signups.event_id
        )
    ).filter(
        Tournament_Signups.tournament_id == tournament_id,
        Tournament_Signups.is_going == True,
        User_Event.active == True
    ).all()
    
    event_dict = {}
    for signup in signups:
        if signup.event_id not in event_dict:
            event_dict[signup.event_id] = []
        event_dict[signup.event_id].append(signup)
    return event_dict

# Helper: Rank signups in each event by weighted points
def rank_signups(event_dict):
    """Rank signups within each event by weighted points (highest first).
    
    Sorts competitors within each event using weighted_points from the metrics system.
    Falls back to user.points or 0 if weighted_points unavailable.
    
    Args:
        event_dict (dict): {event_id: [Tournament_Signups, ...]} from get_signups_by_event.
    
    Returns:
        dict: {event_id: [Tournament_Signups, ...]} - signups sorted by weighted_points descending.
    
    Note:
        Uses getattr with fallback chain: weighted_points → points → 0
    """
    ranked = {}
    for event_id, signups in event_dict.items():
        # Get weighted points from user, fallback to user.points or 0
        ranked[event_id] = sorted(signups, key=lambda s: getattr(s.user, 'weighted_points', getattr(s.user, 'points', 0)), reverse=True)
    return ranked

# Helper: Filter out users with drop penalties and track them
def filter_drops_and_track_penalties(ranked):
    """Filter users with drop penalties and track penalty applications.
    
    Removes users with drops > 0 from rankings, decrements their drop count by 1,
    and tracks penalty information for reporting.
    
    Drop Penalty System:
        - User with drops > 0: Excluded from roster selection
        - User's drops count: Decremented by 1 (penalty applied)
        - Replacement: Next user in ranking without drops
        - Tracking: All penalties logged with original rank and replacement
    
    Args:
        ranked (dict): {event_id: [Tournament_Signups, ...]} sorted by weighted_points.
    
    Returns:
        tuple: (filtered_ranked, penalty_info)
            - filtered_ranked (dict): {event_id: [signups]} without dropped users
            - penalty_info (dict): {event_id: [penalty_dicts]} with penalty details:
                * user_id: Penalized user's ID
                * original_rank: User's rank before exclusion
                * replacement_user_id: Next user without drops (or None)
                * drops: Number of drops user had before decrement
    
    Side Effects:
        - Decrements User.drops for each penalized user (db.session.commit called)
    
    Note:
        This function modifies the database (drops count) and should only be called
        during roster generation, not during preview/view operations.
    """
    filtered_ranked = {}
    penalty_info = {}  # {event_id: [(user_id, rank, replacement_user_id), ...]}
    
    for event_id, signups in ranked.items():
        filtered_signups = []
        penalties = []
        
        for i, signup in enumerate(signups):
            if signup.user.drops > 0:
                # Find replacement (next user without drops)
                replacement = None
                for j in range(i + 1, len(signups)):
                    if signups[j].user.drops == 0:
                        replacement = signups[j].user_id
                        break
                        
                penalties.append({
                    'user_id': signup.user_id,
                    'original_rank': i + 1,
                    'replacement_user_id': replacement,
                    'drops': signup.user.drops
                })
                # Decrement user's drops by 1 (penalty applied)
                signup.user.drops -= 1
                db.session.commit()
            else:
                filtered_signups.append(signup)
        
        filtered_ranked[event_id] = filtered_signups
        if penalties:
            penalty_info[event_id] = penalties
    
    return filtered_ranked, penalty_info

def select_competitors_by_event_type(ranked, speech_spots, ld_spots, pf_spots, event_type_map, judge_children_ids=None, seed_randomness=True, spots_per_event=None):
    """Select competitors using event-type-specific algorithms.
    
    Implements three different selection strategies based on event type:
    
    Speech Events (type 0) - Rotation with Randomness:
        - Rotate through speech events selecting top-ranked competitors
        - If 4+ speech judges: Every 5th selection is random from middle third
        - Ensures fair distribution across speech events
        - Randomness adds variety and gives mid-tier competitors opportunities
    
    Lincoln-Douglas Events (type 1) - Top-N or Middle:
        - If 1 judge: Select top 2 competitors by rank
        - If 2+ judges: Select from middle of rankings (fairness)
        - Middle selection prevents same top competitors dominating
    
    Public Forum Events (type 2) - Top-N or Middle:
        - If 1 judge: Select top 4 competitors by rank
        - If 2+ judges: Select from middle of rankings (fairness)
        - Similar fairness logic as LD
    
    Args:
        ranked (dict): {event_id: [Tournament_Signups, ...]} sorted by weighted_points.
        speech_spots (int): Total speech competitor slots from judge commitments.
        ld_spots (int): Total LD competitor slots.
        pf_spots (int): Total PF competitor slots.
        event_type_map (dict): {event_id: event_type} - 0=Speech, 1=LD, 2=PF.
        seed_randomness (bool): Whether to use random selections for speech (default: True).
    
    Returns:
        tuple: (event_view, rank_view, random_selections)
            - event_view (list): [{'user_id': int, 'event_id': int}, ...] selected competitors
            - rank_view (list): [{'user_id': int, 'event_id': int, 'rank': int}, ...] with rankings
            - random_selections (set): {(user_id, event_id), ...} randomly selected competitors
    
    Algorithm Details:
        Speech Rotation:
            - Cycles through speech events in order
            - Maintains index counter per event
            - Every 5th overall selection: pick random from middle third (if 4+ judges)
            - Stops when all spots filled or all events exhausted
        
        LD/PF Selection:
            - If 2+ judges: middle_index = len(competitors) // 2, select from middle
            - Otherwise: select top N by rank
        
        Random Selection (Speech):
            - Condition: 4+ speech judges AND position % 5 == 4 AND 2+ competitors in event
            - Range: middle third (start = len//3, end = 2*len//3)
            - Tracked in random_selections set for transparency
    
    Note:
        Randomness provides opportunities for mid-tier competitors and prevents
        roster predictability. Middle selection (LD/PF) ensures fairness when
        multiple judges are available.
    """
    event_view = []
    rank_view = []
    random_selections = set()
    selected_user_ids = set()
    if judge_children_ids is None:
        judge_children_ids = set()
    if spots_per_event is None:
        spots_per_event = {}
    
    partnership_map = {}
    for event_id, signups in ranked.items():
        for signup in signups:
            if hasattr(signup, 'partner_id') and signup.partner_id:
                partnership_map[signup.user_id] = signup.partner_id
                partnership_map[signup.partner_id] = signup.user_id
    
    def add_competitor(signup, eid, rank):
        user_id = signup.user_id
        if user_id in selected_user_ids:
            return False
        
        partner_id = partnership_map.get(user_id)
        if partner_id:
            if partner_id in selected_user_ids:
                return False
            partner_in_list = any(s.user_id == partner_id for s in ranked.get(eid, []))
            if not partner_in_list:
                return False
            selected_user_ids.add(partner_id)
        
        selected_user_ids.add(user_id)
        event_view.append({'user_id': user_id, 'event_id': eid})
        rank_view.append({'user_id': user_id, 'event_id': eid, 'rank': rank})
        
        if partner_id:
            event_view.append({'user_id': partner_id, 'event_id': eid})
            partner_rank_info = next((i+1 for i, s in enumerate(ranked.get(eid, [])) if s.user_id == partner_id), rank)
            rank_view.append({'user_id': partner_id, 'event_id': eid, 'rank': partner_rank_info})
        
        return True
    
    speech_event_ids = [eid for eid, etype in event_type_map.items() if etype == 0]
    speech_judges_count = len(speech_event_ids)
    
    for eid in speech_event_ids:
        competitors = ranked.get(eid, [])
        event_max = spots_per_event.get(eid, speech_spots)
        for signup in competitors:
            if signup.user_id in judge_children_ids:
                current_filled = len([e for e in event_view if e['event_id'] == eid])
                if current_filled < event_max:
                    add_competitor(signup, eid, competitors.index(signup) + 1)
    
    speech_indices = {eid: 0 for eid in speech_event_ids}
    speech_filled = len([e for e in event_view if e['event_id'] in speech_event_ids])
    random_counter = 0
    
    while speech_filled < speech_spots and speech_event_ids:
        for eid in speech_event_ids:
            competitors = ranked.get(eid, [])
            if speech_indices[eid] < len(competitors):
                should_be_random = (speech_judges_count >= 4 and random_counter % 5 == 4)
                
                if should_be_random and len(competitors) > 2:
                    mid_start = len(competitors) // 3
                    mid_end = 2 * len(competitors) // 3
                    idx = random.randint(mid_start, mid_end)
                    random_selections.add((competitors[idx].user_id, eid))
                else:
                    idx = speech_indices[eid]
                
                while idx < len(competitors):
                    signup = competitors[idx]
                    if add_competitor(signup, eid, idx + 1):
                        speech_indices[eid] = idx + 1
                        speech_filled = len([e for e in event_view if e['event_id'] in speech_event_ids])
                        random_counter += 1
                        break
                    idx += 1
                    speech_indices[eid] = idx
                
                if speech_filled >= speech_spots:
                    break
        
        if all(speech_indices[eid] >= len(ranked.get(eid, [])) for eid in speech_event_ids):
            break
    
    ld_event_ids = [eid for eid, etype in event_type_map.items() if etype == 1]
    ld_judges_count = len(ld_event_ids)
    
    for eid in ld_event_ids:
        competitors = ranked.get(eid, [])
        event_max = spots_per_event.get(eid, ld_spots)
        for signup in competitors:
            if signup.user_id in judge_children_ids:
                current_filled = len([e for e in event_view if e['event_id'] == eid])
                if current_filled < event_max:
                    add_competitor(signup, eid, competitors.index(signup) + 1)
    
    for eid in ld_event_ids:
        competitors = ranked.get(eid, [])
        event_max = spots_per_event.get(eid, ld_spots)
        filled = len([e for e in event_view if e['event_id'] == eid])
        
        idx = 0
        while filled < event_max and idx < len(competitors):
            if ld_judges_count >= 2 and len(competitors) > 2:
                mid_idx = len(competitors) // 2
                calc_idx = min(idx + mid_idx, len(competitors) - 1)
                random_selections.add((competitors[calc_idx].user_id, eid))
            else:
                calc_idx = idx
            
            attempt = 0
            search_idx = calc_idx
            while search_idx < len(competitors) and attempt < len(competitors):
                signup = competitors[search_idx]
                if add_competitor(signup, eid, search_idx + 1):
                    filled = len([e for e in event_view if e['event_id'] == eid])
                    break
                search_idx += 1
                attempt += 1
            
            idx += 1
            if attempt >= len(competitors):
                break
    
    pf_event_ids = [eid for eid, etype in event_type_map.items() if etype == 2]
    pf_judges_count = len(pf_event_ids)
    
    for eid in pf_event_ids:
        competitors = ranked.get(eid, [])
        event_max = spots_per_event.get(eid, pf_spots)
        for signup in competitors:
            if signup.user_id in judge_children_ids:
                current_filled = len([e for e in event_view if e['event_id'] == eid])
                if current_filled < event_max:
                    add_competitor(signup, eid, competitors.index(signup) + 1)
    
    for eid in pf_event_ids:
        competitors = ranked.get(eid, [])
        event_max = spots_per_event.get(eid, pf_spots)
        filled = len([e for e in event_view if e['event_id'] == eid])
        
        idx = 0
        while filled < event_max and idx < len(competitors):
            if pf_judges_count >= 2 and len(competitors) > 2:
                mid_idx = len(competitors) // 2
                calc_idx = min(idx + mid_idx, len(competitors) - 1)
                random_selections.add((competitors[calc_idx].user_id, eid))
            else:
                calc_idx = idx
            
            attempt = 0
            search_idx = calc_idx
            while search_idx < len(competitors) and attempt < len(competitors):
                signup = competitors[search_idx]
                if add_competitor(signup, eid, search_idx + 1):
                    filled = len([e for e in event_view if e['event_id'] == eid])
                    break
                search_idx += 1
                attempt += 1
            
            idx += 1
            if attempt >= len(competitors):
                break
    
    return event_view, rank_view, random_selections

@rosters_bp.route('/view_tournament/<int:tournament_id>')
def view_tournament(tournament_id):
    """Generate and preview roster for a tournament (live generation, not saved).
    
    Executes the full roster generation algorithm:
    1. Calculate competitor slots from judge commitments
    2. Retrieve and rank all signups by weighted points
    3. Apply drop penalties (filter and track)
    4. Select competitors using event-type-specific algorithms
    5. Display preview with judges, competitors by event, rankings, and penalty info
    
    URL Parameters:
        tournament_id (int): Tournament primary key.
    
    Displays:
        - Judges table: Judge name, child, category (Speech/LD/PF), people bringing
        - Event view: Competitors grouped by event with rankings
        - Rank view: All selected competitors with rank, weighted points, event
        - Penalty info: Users excluded due to drops, replacements, penalty details
        - Random selections: Competitors selected via randomness (speech events)
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Template:
        rosters/view_tournament.html with comprehensive roster preview.
    
    Note:
        This is a PREVIEW - roster is generated dynamically but NOT saved.
        Use save_tournament() to persist the roster.
        Drop penalties ARE applied and decremented during preview.
    """

    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))
    
    speech_competitors, LD_competitors, PF_competitors, spots_per_event = get_roster_count(tournament_id)

    event_dict = get_signups_by_event(tournament_id)
    ranked = rank_signups(event_dict)
    
    filtered_ranked, penalty_info = filter_drops_and_track_penalties(ranked)
    
    event_type_map = {}
    for eid in event_dict.keys():
        event = Event.query.filter_by(id=eid).first()
        if event:
            event_type_map[eid] = event.event_type
    
    judges = Tournament_Judges.query.filter_by(tournament_id=tournament_id, accepted=True).all()
    judge_children_ids = set(j.child_id for j in judges if j.child_id)
    
    event_view, rank_view, random_selections = select_competitors_by_event_type(
        filtered_ranked,
        speech_spots=speech_competitors,
        ld_spots=LD_competitors,
        pf_spots=PF_competitors,
        event_type_map=event_type_map,
        judge_children_ids=judge_children_ids,
        seed_randomness=True,
        spots_per_event=spots_per_event
    )

    # Build event_competitors: {event_id: [comp, ...]} with rank info
    event_competitors = {}
    for comp in event_view:
        eid = comp['event_id']
        if eid not in event_competitors:
            event_competitors[eid] = []
        # Find the corresponding rank from rank_view
        rank_info = next((r for r in rank_view if r['user_id'] == comp['user_id'] and r['event_id'] == comp['event_id']), None)
        comp_with_rank = comp.copy()
        comp_with_rank['rank'] = rank_info['rank'] if rank_info else 'N/A'
        event_competitors[eid].append(comp_with_rank)

    # Build users and events dicts for template lookup
    user_ids = set([comp['user_id'] for comp in event_view] + [row['user_id'] for row in rank_view])
    event_ids = set([comp['event_id'] for comp in event_view] + [row['event_id'] for row in rank_view])

    users = {}
    events = {}

    if user_ids:
        users = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()}
    if event_ids:
        events = {e.id: e for e in Event.query.filter(Event.id.in_(event_ids)).all()}

    # Get point weights for weighted points calculation
    tournament_weight, effort_weight = get_point_weights()

    for eid, competitors in event_competitors.items():
        for comp in competitors:
            user = users.get(comp['user_id'])
            if user:
                comp['weighted_points'] = calculate_weighted_points(user)

    # Judges for the tournament
    judges = Tournament_Judges.query.filter_by(tournament_id=tournament_id, accepted=True).all()
    judge_user_ids = [j.judge_id for j in judges if j.judge_id]
    child_user_ids = [j.child_id for j in judges if j.child_id]
    all_judge_user_ids = list(set(judge_user_ids + child_user_ids))
    
    judge_users = {}
    if all_judge_user_ids:
        judge_users = {u.id: u for u in User.query.filter(User.id.in_(all_judge_user_ids)).all()}

    # Debug output
    print(f"Tournament {tournament_id}: {len(judges)} judges, {len(event_view)} competitors in event_view, {len(rank_view)} in rank_view")
    print(f"Event competitors: {list(event_competitors.keys())}")
    print(f"Users dict has {len(users)} users")
    print(f"Events dict has {len(events)} events")

    tournament = Tournament.query.get(tournament_id)
    return render_template('rosters/view_tournament.html',
                          tournament=tournament,
                          event_view=event_view,
                          rank_view=rank_view,
                          event_competitors=event_competitors,
                          users=users,
                          events=events,
                          judges=judges,
                          judge_users=judge_users,
                          penalty_info=penalty_info,
                          random_selections=random_selections,
                          tournament_weight=tournament_weight,
                          effort_weight=effort_weight,
                          upcoming_tournaments=[],
                          tournaments=[],
                          rosters=[])

@rosters_bp.route('/download_tournament/<int:tournament_id>')
def download_tournament(tournament_id):
    """Export live-generated tournament roster to Excel with multiple sheets.
    
    Generates roster using same algorithm as view_tournament() and exports to Excel.
    Uses pandas and openpyxl for Excel creation (requires optional dependencies).
    
    URL Parameters:
        tournament_id (int): Tournament primary key.
    
    Excel Format:
        Sheet 1 - Judges:
            Columns: Judge Name, Child, Category, Number People Bringing, Judge ID,
                    Child ID, Event ID
        
        Sheet 2 - Rank View:
            Columns: Rank, Competitor Name, Weighted Points, Event, Category,
                    User ID, Event ID
            Sorted by rank within each event
        
        Sheet 3+ - Event Sheets (one per event):
            Columns: Event, Category, Rank, Competitor, Weighted Points,
                    User ID, Event ID
            Sheet name: First 30 chars of event name (sanitized)
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        Excel file download (.xlsx): tournament_{name}_{timestamp}.xlsx
        Redirects to view_tournament if pandas/openpyxl unavailable.
    
    Dependencies:
        - pandas: DataFrame operations
        - openpyxl: Excel file writing
        Falls back gracefully if not installed.
    
    Note:
        This generates a fresh roster (not from saved Roster record).
        Use download_roster() for saved/published rosters.
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    tournament = Tournament.query.get(tournament_id)
    if not tournament:
        flash("Tournament not found")
        return redirect(url_for('rosters.index'))

    if pd is None or openpyxl is None:
        flash("Excel functionality not available. Please install pandas and openpyxl.")
        return redirect(url_for('rosters.view_tournament', tournament_id=tournament_id))

    speech_competitors, LD_competitors, PF_competitors, spots_per_event = get_roster_count(tournament_id)
    event_dict = get_signups_by_event(tournament_id)
    ranked = rank_signups(event_dict)
    event_type_map = {}
    for eid in event_dict.keys():
        event = Event.query.filter_by(id=eid).first()
        if event:
            event_type_map[eid] = event.event_type
    
    judges = Tournament_Judges.query.filter_by(tournament_id=tournament_id, accepted=True).all()
    judge_children_ids = set(j.child_id for j in judges if j.child_id)
    
    event_view, rank_view, random_selections = select_competitors_by_event_type(
        ranked,
        speech_spots=speech_competitors,
        ld_spots=LD_competitors,
        pf_spots=PF_competitors,
        event_type_map=event_type_map,
        judge_children_ids=judge_children_ids,
        seed_randomness=True,
        spots_per_event=spots_per_event
    )

    # Build event_competitors: {event_id: [comp, ...]} with rank info
    event_competitors = {}
    for comp in event_view:
        eid = comp['event_id']
        if eid not in event_competitors:
            event_competitors[eid] = []
        # Find the corresponding rank from rank_view
        rank_info = next((r for r in rank_view if r['user_id'] == comp['user_id'] and r['event_id'] == comp['event_id']), None)
        comp_with_rank = comp.copy()
        comp_with_rank['rank'] = rank_info['rank'] if rank_info else 'N/A'
        comp_with_rank['is_random'] = (comp['user_id'], comp['event_id']) in random_selections
        event_competitors[eid].append(comp_with_rank)

    # Build users and events dicts for template lookup
    user_ids = set([comp['user_id'] for comp in event_view] + [row['user_id'] for row in rank_view])
    event_ids = set([comp['event_id'] for comp in event_view] + [row['event_id'] for row in rank_view])
    users = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()}
    events = {e.id: e for e in Event.query.filter(Event.id.in_(event_ids)).all()}

    # Judges for the tournament
    judges = Tournament_Judges.query.filter_by(tournament_id=tournament_id, accepted=True).all()
    judge_user_ids = [j.judge_id for j in judges if j.judge_id]
    child_user_ids = [j.child_id for j in judges if j.child_id]
    all_judge_user_ids = list(set(judge_user_ids + child_user_ids))
    judge_users = {u.id: u for u in User.query.filter(User.id.in_(all_judge_user_ids)).all() if all_judge_user_ids}

    # Create Excel file with multiple sheets
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Judges sheet
    judges_data = []
    for judge in judges:
        judge_name = f"{judge_users[judge.judge_id].first_name} {judge_users[judge.judge_id].last_name}" if judge.judge_id in judge_users else 'Unknown'
        child_name = f"{judge.child.first_name} {judge.child.last_name}" if judge.child else ''
        event_type = 'Unknown'
        people_bringing = 0
        if judge.event:
            if judge.event.event_type == 0:
                event_type = 'Speech'
                people_bringing = 6
            elif judge.event.event_type == 1:
                event_type = 'LD'
                people_bringing = 2
            elif judge.event.event_type == 2:
                event_type = 'PF'
                people_bringing = 4
        
        judges_data.append({
            'Judge Name': judge_name,
            'Child': child_name,
            'Category': event_type,
            'Number People Bringing': people_bringing,
            'Judge ID': judge.judge_id,
            'Child ID': judge.child_id,
            'Event ID': judge.event_id
        })
    
    judges_df = pd.DataFrame(judges_data)
    judges_df.to_excel(writer, sheet_name='Judges', index=False)

    # Rank View sheet
    rank_data = []
    for row in rank_view:
        user = users.get(row['user_id'])
        user_name = f"{user.first_name} {user.last_name}" if user else 'Unknown'
        event_name = events[row['event_id']].event_name if row['event_id'] in events else 'Unknown Event'
        event_type = 'Unknown'
        if row['event_id'] in events:
            if events[row['event_id']].event_type == 0:
                event_type = 'Speech'
            elif events[row['event_id']].event_type == 1:
                event_type = 'LD'
            elif events[row['event_id']].event_type == 2:
                event_type = 'PF'
        weighted_points = calculate_weighted_points(user)
        rank_data.append({
            'Rank': row['rank'],
            'Competitor Name': user_name,
            'Weighted Points': weighted_points,
            'Event': event_name,
            'Category': event_type,
            'User ID': row['user_id'],
            'Event ID': row['event_id']
        })

    rank_df = pd.DataFrame(rank_data)
    rank_df.to_excel(writer, sheet_name='Rank View', index=False)

    # Event View sheets (one for each event)
    for event_id, competitors_list in event_competitors.items():
        event_name = events[event_id].event_name if event_id in events else f'Event {event_id}'
        event_data = []

        for comp in competitors_list:
            user = users.get(comp['user_id'])
            user_name = f"{user.first_name} {user.last_name}" if user else 'Unknown'
            event_type = 'Unknown'
            if event_id in events:
                if events[event_id].event_type == 0:
                    event_type = 'Speech'
                elif events[event_id].event_type == 1:
                    event_type = 'LD'
                elif events[event_id].event_type == 2:
                    event_type = 'PF'
            weighted_points = calculate_weighted_points(user)
            event_data.append({
                'Event': event_name,
                'Category': event_type,
                'Rank': comp['rank'],
                'Competitor': user_name,
                'Weighted Points': weighted_points,
                'User ID': comp['user_id'],
                'Event ID': comp['event_id']
            })

        event_df = pd.DataFrame(event_data)
        # Limit sheet name length and remove invalid characters
        sheet_name = event_name[:30].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace(':', '-').replace('[', '-').replace(']', '-')
        event_df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()
    output.seek(0)

    filename = f"tournament_{tournament.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, 
                     as_attachment=True, 
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@rosters_bp.route('/save_roster/<int:tournament_id>')
def save_tournament(tournament_id):
    """Save generated roster to database for permanent storage.
    
    Executes roster generation algorithm and persists results to database:
    1. Generate roster using same algorithm as view_tournament()
    2. Create Roster record with timestamp
    3. Save all Roster_Competitors entries
    4. Save partner relationships (for partner events like PF)
    5. Save Roster_Judge entries with people_bringing calculations
    6. Save penalty entries for tracking drop applications
    
    URL Parameters:
        tournament_id (int): Tournament primary key.
    
    Database Records Created:
        - Roster: Main roster record with name, date, tournament_id
        - Roster_Competitors: One per selected competitor (user_id, event_id, roster_id)
        - Roster_Partners: Partner pairs for partner events (both users selected)
        - Roster_Judge: One per judge (user_id, child_id, event_id, people_bringing)
        - Roster_Penalty_Entries: Drop penalties applied (user, rank, drops, event)
    
    Roster Naming:
        Format: "{tournament.name} {timestamp}" in US/Eastern timezone
        Example: "Harvard Tournament 2025-10-11 14:30:45"
    
    Partner Event Handling:
        - Checks if event.is_partner_event is True
        - Retrieves partner_id from Tournament_Signups
        - Only creates partnership if BOTH partners selected for roster
        - Avoids duplicate partnerships using processed_partnerships set
        - Stores sorted pair (lower ID first) in Roster_Partners
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        Redirects to view_roster with new roster_id.
    
    Side Effects:
        - Creates multiple database records
        - Applies drop penalties (decrements User.drops)
    
    Note:
        This is the permanent save operation. Use view_tournament() for preview.
        Drop penalties are applied during save and persist in database.
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    speech_competitors, LD_competitors, PF_competitors, spots_per_event = get_roster_count(tournament_id)
    event_dict = get_signups_by_event(tournament_id)
    ranked = rank_signups(event_dict)
    
    filtered_ranked, penalty_info = filter_drops_and_track_penalties(ranked)
    
    event_type_map = {}
    for eid in event_dict.keys():
        event = Event.query.filter_by(id=eid).first()
        if event:
            event_type_map[eid] = event.event_type
    
    judges = Tournament_Judges.query.filter_by(tournament_id=tournament_id, accepted=True).all()
    judge_children_ids = set(j.child_id for j in judges if j.child_id)
    
    event_view, rank_view, random_selections = select_competitors_by_event_type(
        filtered_ranked,
        speech_spots=speech_competitors,
        ld_spots=LD_competitors,
        pf_spots=PF_competitors,
        event_type_map=event_type_map,
        judge_children_ids=judge_children_ids,
        seed_randomness=True,
        spots_per_event=spots_per_event
    )

    tz = pytz.timezone('US/Eastern')
    tournament = Tournament.query.get(tournament_id)
    roster_name = f"{tournament.name} {datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')}"
    new_roster = Roster(name=roster_name, date_made=datetime.now(tz), tournament_id=tournament_id)
    db.session.add(new_roster)
    db.session.commit()  # Commit to get the roster id

    # Save competitors using the event_view generated by the helpers
    for comp in event_view:
        rc = Roster_Competitors(
            user_id=comp['user_id'],
            event_id=comp['event_id'],
            judge_id=None,  # Optionally, could be filled if logic exists
            roster_id=new_roster.id
        )
        db.session.add(rc)

    # Save partner relationships for partner events
    from mason_snd.models.rosters import Roster_Partners
    processed_partnerships = set()  # To avoid duplicate partnerships
    
    for comp in event_view:
        # Check if this is a partner event
        event = Event.query.get(comp['event_id'])
        if event and event.is_partner_event:
            # Find the signup to get partner information
            signup = Tournament_Signups.query.join(
                User_Event,
                db.and_(
                    User_Event.user_id == Tournament_Signups.user_id,
                    User_Event.event_id == Tournament_Signups.event_id
                )
            ).filter(
                Tournament_Signups.tournament_id == tournament_id,
                Tournament_Signups.user_id == comp['user_id'],
                Tournament_Signups.event_id == comp['event_id'],
                Tournament_Signups.is_going == True,
                User_Event.active == True
            ).first()
            
            if signup and signup.partner_id:
                # Check if partner is also selected for the roster
                partner_in_roster = any(
                    c['user_id'] == signup.partner_id and c['event_id'] == comp['event_id'] 
                    for c in event_view
                )
                
                if partner_in_roster:
                    # Create partnership entry (avoid duplicates)
                    partnership_key = tuple(sorted([comp['user_id'], signup.partner_id]))
                    if partnership_key not in processed_partnerships:
                        rp = Roster_Partners(
                            partner1_user_id=partnership_key[0],
                            partner2_user_id=partnership_key[1],
                            roster_id=new_roster.id
                        )
                        db.session.add(rp)
                        processed_partnerships.add(partnership_key)

    # Save judges using the current Tournament_Judges with proper people_bringing calculation
    judges = Tournament_Judges.query.filter_by(tournament_id=tournament_id, accepted=True).all()
    for judge in judges:
        # Calculate people_bringing based on event type
        people_bringing = 0
        if judge.event:
            if judge.event.event_type == 0:  # Speech
                people_bringing = 6
            elif judge.event.event_type == 1:  # LD
                people_bringing = 2
            elif judge.event.event_type == 2:  # PF
                people_bringing = 4
        
        rj = Roster_Judge(
            user_id=judge.judge_id,
            child_id=judge.child_id,
            event_id=judge.event_id,
            roster_id=new_roster.id,
            people_bringing=people_bringing
        )
        db.session.add(rj)

    # Save penalty entries
    for event_id, penalties in penalty_info.items():
        for penalty in penalties:
            rpe = Roster_Penalty_Entries(
                roster_id=new_roster.id,
                tournament_id=tournament_id,
                event_id=event_id,
                penalized_user_id=penalty['user_id'],
                original_rank=penalty['original_rank'],
                drops_applied=penalty['drops']
            )
            db.session.add(rpe)

    db.session.commit()
    flash("Roster saved!")
    return redirect(url_for('rosters.view_roster', roster_id=new_roster.id))


@rosters_bp.route('/publish_roster/<int:roster_id>')
def publish_roster(roster_id):
    """Publish roster to make it visible to users and create notifications.
    
    Publishing Process:
    1. Mark roster.published = True and set published_at timestamp
    2. Create User_Published_Rosters entries for all competitors
    3. Users can then see roster in their profile
    4. Notification system can alert users (notified flag tracks this)
    
    URL Parameters:
        roster_id (int): Roster primary key.
    
    Database Changes:
        - Roster.published: False → True
        - Roster.published_at: NULL → current timestamp (US/Eastern)
        - User_Published_Rosters: Created for each competitor
            * user_id, roster_id, tournament_id, event_id, notified=False
    
    User_Published_Rosters Purpose:
        - Links users to published rosters they're on
        - Enables "My Rosters" view in user profiles
        - Tracks notification status (notified field)
        - Allows users to see tournament/event assignments
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        Redirects to view_roster with success message.
        Flash error if roster already published.
    
    Note:
        Publishing is irreversible without unpublish_roster().
        Users are notified of roster assignments via User_Published_Rosters.
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    roster = Roster.query.get_or_404(roster_id)
    
    if roster.published:
        flash("This roster is already published!")
        return redirect(url_for('rosters.view_roster', roster_id=roster_id))

    # Mark roster as published
    tz = pytz.timezone('US/Eastern')
    roster.published = True
    roster.published_at = datetime.now(tz)
    
    # Create entries for all users on this roster so they can see it in their profile
    competitors = Roster_Competitors.query.filter_by(roster_id=roster_id).all()
    for competitor in competitors:
        # Check if entry already exists
        existing = User_Published_Rosters.query.filter_by(
            user_id=competitor.user_id,
            roster_id=roster_id
        ).first()
        
        if not existing:
            published_entry = User_Published_Rosters(
                user_id=competitor.user_id,
                roster_id=roster_id,
                tournament_id=roster.tournament_id,
                event_id=competitor.event_id,
                notified=False
            )
            db.session.add(published_entry)
    
    db.session.commit()
    flash("Roster has been published! Users will be notified.")
    return redirect(url_for('rosters.view_roster', roster_id=roster_id))


@rosters_bp.route('/unpublish_roster/<int:roster_id>')
def unpublish_roster(roster_id):
    """Unpublish roster to hide it from users (reverses publish_roster).
    
    Unpublishing Process:
    1. Mark roster.published = False and clear published_at timestamp
    2. Delete all User_Published_Rosters entries for this roster
    3. Users can no longer see roster in their profiles
    4. Roster remains in database but is hidden
    
    URL Parameters:
        roster_id (int): Roster primary key.
    
    Database Changes:
        - Roster.published: True → False
        - Roster.published_at: timestamp → NULL
        - User_Published_Rosters: All entries deleted (cascade)
    
    Use Cases:
        - Roster needs corrections after publishing
        - Tournament cancelled or rescheduled
        - Mistake in roster generation
        - Want to regenerate roster with updated data
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        Redirects to view_roster with confirmation message.
    
    Note:
        Unpublishing does NOT delete the roster, only hides it from users.
        Use delete_roster() to permanently remove.
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    roster = Roster.query.get_or_404(roster_id)
    
    # Mark roster as unpublished
    roster.published = False
    roster.published_at = None
    
    # Remove published roster entries for users
    User_Published_Rosters.query.filter_by(roster_id=roster_id).delete()
    
    db.session.commit()
    flash("Roster has been unpublished.")
    return redirect(url_for('rosters.view_roster', roster_id=roster_id))



@rosters_bp.route('/view_roster/<int:roster_id>')
def view_roster(roster_id):
    """Display a saved roster from database (not live-generated).
    
    Retrieves roster data from Roster, Roster_Competitors, Roster_Judge, and
    Roster_Penalty_Entries tables and displays using same layout as view_tournament.
    
    URL Parameters:
        roster_id (int): Roster primary key.
    
    Data Retrieved:
        - Roster: Basic info (name, date_made, tournament_id, published status)
        - Roster_Competitors: All selected competitors (user_id, event_id)
        - Roster_Judge: All judges (user_id, child_id, event_id, people_bringing)
        - Roster_Penalty_Entries: Drop penalties applied during generation
    
    Displays:
        - Roster metadata (name, date, published status)
        - Judges table with child and event assignments
        - Event view (competitors grouped by event)
        - Rank view (all competitors with rankings and weighted points)
        - Penalty info (users excluded, original ranks, drops applied)
    
    Ranking Logic:
        Since saved rosters don't recalculate rankings, uses index + 1 as rank.
        For display purposes, competitors are shown in order they were saved.
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Template:
        rosters/view_roster.html with saved roster data.
    
    Note:
        This displays SAVED data, not live-generated. Changes to signups or
        weighted points after save are NOT reflected.
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    roster = Roster.query.get(roster_id)
    if not roster:
        flash("Roster not found")
        return redirect(url_for('rosters.index'))

    # Get competitors and judges from the roster
    competitors = Roster_Competitors.query.filter_by(roster_id=roster_id).all()
    judges = Roster_Judge.query.filter_by(roster_id=roster_id).all()
    
    # Get penalty entries for this roster
    penalty_entries = Roster_Penalty_Entries.query.filter_by(roster_id=roster_id).all()
    penalty_info = {}
    for entry in penalty_entries:
        if entry.event_id not in penalty_info:
            penalty_info[entry.event_id] = []
        penalty_info[entry.event_id].append({
            'user_id': entry.penalized_user_id,
            'original_rank': entry.original_rank,
            'drops': entry.drops_applied,
            'replacement_user_id': None  # For saved rosters, we don't track replacements
        })

    # Build event_view and rank_view from competitors
    event_view = []
    rank_view = []
    event_competitors = {}
    
    for i, comp in enumerate(competitors):
        event_view.append({'user_id': comp.user_id, 'event_id': comp.event_id})
        # Use index + 1 as rank for saved rosters
        rank_view.append({'user_id': comp.user_id, 'event_id': comp.event_id, 'rank': i + 1})
        
        eid = comp.event_id
        if eid not in event_competitors:
            event_competitors[eid] = []
        event_competitors[eid].append({'user_id': comp.user_id, 'event_id': eid, 'rank': i + 1})

    # Build users and events dicts for template lookup
    user_ids = set([comp.user_id for comp in competitors] + [j.user_id for j in judges if j.user_id] + [j.child_id for j in judges if j.child_id])
    event_ids = set([comp.event_id for comp in competitors] + [j.event_id for j in judges if j.event_id])
    
    users = {}
    events = {}
    
    if user_ids:
        users = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()}
    if event_ids:
        events = {e.id: e for e in Event.query.filter(Event.id.in_(event_ids)).all()}

    # Get point weights for weighted points calculation
    tournament_weight, effort_weight = get_point_weights()

    # Judges for the roster
    judge_user_ids = [j.user_id for j in judges if j.user_id]
    child_user_ids = [j.child_id for j in judges if j.child_id]
    all_judge_user_ids = list(set(judge_user_ids + child_user_ids))
    
    judge_users = {}
    if all_judge_user_ids:
        judge_users = {u.id: u for u in User.query.filter(User.id.in_(all_judge_user_ids)).all()}

    # Debug information
    print(f"Roster {roster_id}: {len(competitors)} competitors, {len(judges)} judges")
    print(f"Event view has {len(event_view)} entries")
    print(f"Rank view has {len(rank_view)} entries")
    print(f"Event competitors: {list(event_competitors.keys())}")
    print(f"Users dict has {len(users)} users")
    print(f"Events dict has {len(events)} events")

    return render_template('rosters/view_roster.html',
                          roster=roster,
                          event_view=event_view,
                          rank_view=rank_view,
                          event_competitors=event_competitors,
                          users=users,
                          events=events,
                          judges=judges,
                          judge_users=judge_users,
                          penalty_info=penalty_info,
                          tournament_weight=tournament_weight,
                          effort_weight=effort_weight,
                          upcoming_tournaments=[],
                          tournaments=[],
                          rosters=[])


# AJAX endpoint: search judges (parents) for autocomplete on view_roster
@rosters_bp.route('/search_judges')
def search_judges():
    from flask import jsonify
    q = request.args.get('q', '').strip()
    roster_id = request.args.get('roster_id')

    if not q or len(q) < 1:
        return jsonify({'users': []})

    # Search for users who are parents (is_parent=True) matching name
    # Include ghost accounts (account_claimed=False) so admins can add placeholders
    users = User.query.filter(
        db.and_(
            User.is_parent == True,
            db.or_(
                User.first_name.ilike(f'%{q}%'),
                User.last_name.ilike(f'%{q}%'),
                db.func.concat(User.first_name, ' ', User.last_name).ilike(f'%{q}%')
            )
        )
    ).limit(20).all()

    return jsonify({'users': [{'id': u.id, 'first_name': u.first_name, 'last_name': u.last_name} for u in users]})


# AJAX endpoint: get children for a specific judge
@rosters_bp.route('/get_judge_children')
def get_judge_children():
    from flask import jsonify
    from mason_snd.models.auth import Judges as JudgesRelationship
    
    judge_id = request.args.get('judge_id')
    
    if not judge_id:
        return jsonify({'children': []})
    
    try:
        judge_id = int(judge_id)
    except (ValueError, TypeError):
        return jsonify({'children': []})
    
    # Get all children for this judge from Judges relationship table
    relationships = JudgesRelationship.query.filter_by(judge_id=judge_id).all()
    
    children = []
    for rel in relationships:
        if rel.child_id and rel.child:
            children.append({
                'id': rel.child.id,
                'first_name': rel.child.first_name,
                'last_name': rel.child.last_name
            })
    
    return jsonify({'children': children})


def _auto_fill_roster_from_signups(roster_id):
    """Auto-fill roster with additional competitors from tournament signups.
    
    When judges are added and there's additional capacity, this function:
    1. Calculates total judge capacity per event type
    2. Finds how many competitor slots are available
    3. Pulls additional competitors from Tournament_Signups (ranked by weighted_points)
    4. Adds them to fill the available capacity
    
    Only adds competitors if roster has a tournament_id.
    """
    roster = Roster.query.get(roster_id)
    if not roster or not roster.tournament_id:
        return
    
    # Get current judges and their capacities per event
    judges = Roster_Judge.query.filter_by(roster_id=roster_id).all()
    capacity_by_event_type = {0: 0, 1: 0, 2: 0}  # Speech, LD, PF
    
    for judge in judges:
        if judge.event_id:
            event = Event.query.get(judge.event_id)
            if event:
                capacity_by_event_type[event.event_type] += (judge.people_bringing or 0)
    
    # Get current competitors count per event type
    competitors = Roster_Competitors.query.filter_by(roster_id=roster_id).all()
    current_by_event_type = {0: 0, 1: 0, 2: 0}
    
    for comp in competitors:
        if comp.event_id:
            event = Event.query.get(comp.event_id)
            if event:
                current_by_event_type[event.event_type] += 1
    
    # For each event type, add competitors if we have capacity
    for event_type in [0, 1, 2]:
        capacity = capacity_by_event_type[event_type]
        current = current_by_event_type[event_type]
        needed = capacity - current
        
        if needed <= 0:
            continue
        
        # Get events of this type
        events_of_type = Event.query.filter_by(event_type=event_type).all()
        event_ids = [e.id for e in events_of_type]
        
        if not event_ids:
            continue
        
        # Get signups for this tournament and these events, not already in roster
        existing_user_ids = set([c.user_id for c in competitors])
        
        # Get signups ranked by weighted_points (only active users in events)
        signups = Tournament_Signups.query.join(
            User_Event,
            db.and_(
                User_Event.user_id == Tournament_Signups.user_id,
                User_Event.event_id == Tournament_Signups.event_id
            )
        ).filter(
            Tournament_Signups.tournament_id == roster.tournament_id,
            Tournament_Signups.event_id.in_(event_ids),
            Tournament_Signups.is_going == True,
            User_Event.active == True,
            ~Tournament_Signups.user_id.in_(existing_user_ids)
        ).all()
        
        # Sort by weighted_points (highest first)
        signups_sorted = sorted(signups, 
                               key=lambda s: getattr(s.user, 'weighted_points', getattr(s.user, 'points', 0)),
                               reverse=True)
        
        # Add top N signups to fill capacity
        for signup in signups_sorted[:needed]:
            rc = Roster_Competitors(
                user_id=signup.user_id,
                event_id=signup.event_id,
                judge_id=None,
                roster_id=roster_id
            )
            db.session.add(rc)
            print(f"Auto-filled: Added {signup.user.first_name} {signup.user.last_name} to event {signup.event_id}")
    
    db.session.flush()


def _redistribute_competitors_for_roster(roster_id):
    """Simple redistribution algorithm for saved roster when judges change.

    For each event in the roster, assign competitors (in current saved order)
    to available Roster_Judge entries for that event based on people_bringing.
    This preserves ordering while ensuring judge capacities are respected.
    """
    # Load current competitors and judges
    competitors = Roster_Competitors.query.filter_by(roster_id=roster_id).all()
    judges = Roster_Judge.query.filter_by(roster_id=roster_id).all()

    # Group judges by event
    judges_by_event = {}
    for j in judges:
        judges_by_event.setdefault(j.event_id, []).append(j)

    # Clear existing judge assignments
    for comp in competitors:
        comp.judge_id = None
    db.session.flush()

    # Assign per event
    for event_id, comps in {}.items():
        pass

    # Build mapping of event -> competitors (preserve saved order)
    comps_by_event = {}
    for comp in competitors:
        comps_by_event.setdefault(comp.event_id, []).append(comp)

    # For each event, assign sequentially to judges respecting capacity
    for event_id, comps in comps_by_event.items():
        available_judges = judges_by_event.get(event_id, [])
        if not available_judges:
            # No judges for this event; leave judge_id as None
            continue

        # Expand judge slots into a list of judge ids repeated by capacity
        slot_judge_ids = []
        for j in available_judges:
            cap = j.people_bringing or 0
            # If people_bringing is 0, treat as 1 (judge themselves)
            if cap <= 0:
                cap = 1
            # We store judge.user_id as the judge identifier
            slot_judge_ids.extend([j.user_id] * cap)

        # Assign each competitor to the next available judge slot
        slot_index = 0
        for comp in comps:
            if slot_index >= len(slot_judge_ids):
                # No more capacity; leave unassigned
                comp.judge_id = None
            else:
                comp.judge_id = slot_judge_ids[slot_index]
                slot_index += 1

    db.session.commit()


@rosters_bp.route('/add_roster_judge', methods=['POST'])
def add_roster_judge():
    """Add a judge to a saved roster and redistribute competitors.

    Expects form data: roster_id, user_id (judge), child_id (optional), event_id (optional)
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash('Log In First')
        return redirect_to_login()

    if user.role < 2:
        flash('You are not authorized to perform this action')
        return redirect(url_for('main.index'))

    roster_id = request.form.get('roster_id')
    judge_user_id = request.form.get('user_id')
    child_id = request.form.get('child_id')
    event_type = request.form.get('event_type')  # Get event_type directly (0=Speech, 1=LD, 2=PF)

    if not roster_id or not judge_user_id:
        flash('Missing parameters: roster_id or judge user_id')
        return redirect(request.referrer or url_for('rosters.view_roster', roster_id=roster_id))

    if not event_type:
        flash('Missing category selection (Speech/LD/PF)')
        return redirect(request.referrer or url_for('rosters.view_roster', roster_id=roster_id))

    roster = Roster.query.get(roster_id)
    if not roster:
        flash('Roster not found')
        return redirect(request.referrer or url_for('rosters.index'))

    # Determine people_bringing based on event_type
    try:
        event_type_int = int(event_type)
    except (ValueError, TypeError):
        flash('Invalid category')
        return redirect(url_for('rosters.view_roster', roster_id=roster_id))

    if event_type_int == 0:  # Speech
        people_bringing = 6
    elif event_type_int == 1:  # LD
        people_bringing = 2
    elif event_type_int == 2:  # PF
        people_bringing = 4
    else:
        flash('Invalid category type')
        return redirect(url_for('rosters.view_roster', roster_id=roster_id))

    # Find an actual event of this type if possible, or set event_id to None
    # For roster purposes, we may need to find or create a matching event
    # For now, find first event matching the event_type
    event_id = None
    if child_id and roster.tournament_id:
        # Try to find child's signup event first (only if actively enrolled)
        ts = Tournament_Signups.query.join(
            User_Event,
            db.and_(
                User_Event.user_id == Tournament_Signups.user_id,
                User_Event.event_id == Tournament_Signups.event_id
            )
        ).filter(
            Tournament_Signups.user_id == child_id,
            Tournament_Signups.tournament_id == roster.tournament_id,
            Tournament_Signups.is_going == True,
            User_Event.active == True
        ).first()
        if ts and ts.event_id:
            ev = Event.query.get(ts.event_id)
            if ev and ev.event_type == event_type_int:
                event_id = ts.event_id
    
    # If still no event_id, find any event of this type
    if not event_id:
        ev = Event.query.filter_by(event_type=event_type_int).first()
        if ev:
            event_id = ev.id

    # Check for duplicate with same judge/child/event_type (allow multiple if different events)
    # Since we're now using event_type, check if judge already exists with same type
    existing_judges = Roster_Judge.query.filter_by(roster_id=roster_id, user_id=judge_user_id, child_id=child_id).all()
    for existing in existing_judges:
        if existing.event_id:
            existing_event = Event.query.get(existing.event_id)
            if existing_event and existing_event.event_type == event_type_int:
                flash('Judge already exists for this roster with this category')
                return redirect(url_for('rosters.view_roster', roster_id=roster_id))

    new_j = Roster_Judge(user_id=judge_user_id, child_id=child_id or None, event_id=event_id, roster_id=roster_id, people_bringing=people_bringing)
    db.session.add(new_j)
    db.session.commit()

    # Auto-fill roster with additional competitors if capacity available
    _auto_fill_roster_from_signups(roster_id)
    
    # Redistribute competitors for roster
    _redistribute_competitors_for_roster(roster_id)

    flash('Judge added, roster auto-filled, and competitors redistributed')
    return redirect(url_for('rosters.view_roster', roster_id=roster_id))


@rosters_bp.route('/remove_roster_judge', methods=['POST'])
def remove_roster_judge():
    """Remove a judge from a saved roster and redistribute competitors.

    Expects form data: roster_id, roster_judge_id
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash('Log In First')
        return redirect_to_login()

    if user.role < 2:
        flash('You are not authorized to perform this action')
        return redirect(url_for('main.index'))

    roster_id = request.form.get('roster_id')
    roster_judge_id = request.form.get('roster_judge_id')

    if not roster_id or not roster_judge_id:
        flash('Missing parameters')
        return redirect(request.referrer or url_for('rosters.view_roster', roster_id=roster_id))

    rj = Roster_Judge.query.filter_by(id=roster_judge_id, roster_id=roster_id).first()
    if not rj:
        flash('Judge entry not found')
        return redirect(url_for('rosters.view_roster', roster_id=roster_id))

    db.session.delete(rj)
    db.session.commit()

    # Redistribute competitors for roster
    _redistribute_competitors_for_roster(roster_id)

    flash('Judge removed and competitors redistributed')
    return redirect(url_for('rosters.view_roster', roster_id=roster_id))

@rosters_bp.route('/download_roster/<int:roster_id>')
def download_roster(roster_id):
    """Export saved roster to Excel - fully editable and re-uploadable.
    
    Creates comprehensive Excel file from saved Roster record with:
    - Multiple sheets (Judges, Rank View, per-event views)
    - Formatted columns (color-coded IDs, editable fields, read-only info)
    - Partner information (for partner events)
    - IDs embedded for smart reconciliation on re-upload
    
    URL Parameters:
        roster_id (int): Roster primary key.
    
    Excel Format:
        Sheet 1 - Judges:
            Columns: Judge Name, Child, Event, Category, Number People Bringing,
                    Judge ID, Child ID, Event ID
            Color coding: ID columns (light blue), editable (white), info (gray)
        
        Sheet 2 - Rank View (PRIMARY EDITING SHEET):
            Columns: Rank, Competitor Name, Partner, Weighted Points, Event,
                    Category, Status, User ID, Partner ID, Event ID
            Sorted by weighted points within each event
            Full partner tracking for PF events
        
        Sheet 3+ - Event Sheets (one per event):
            Columns: Event, Category, Rank, Competitor, Partner, Weighted Points,
                    User ID, Partner ID, Event ID
            Sorted by weighted points per event
    
    Color Coding:
        - Header row: Dark blue with white text
        - ID columns: Light blue (read-only indicators, used for matching)
        - Info columns (Rank, Points, Category): Light gray (calculated/informational)
        - Editable columns (Names, Partner): White (can be modified)
    
    Partner Information:
        - Retrieves Roster_Partners for this roster
        - Builds partnership_map: {user_id: partner_user_id}
        - Displays partner names and IDs in export
        - Enables partner tracking on re-upload
    
    Smart Reconciliation Support:
        - Embeds User ID, Partner ID, Event ID for accurate matching
        - Names take precedence for display, IDs for matching
        - On re-upload: IDs matched first, then fuzzy name matching
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        Excel file download (.xlsx): roster_{name}_{timestamp}.xlsx
        Redirects to view_roster if pandas/openpyxl unavailable.
    
    Dependencies:
        - pandas: DataFrame creation
        - openpyxl: Excel writing and formatting
        Falls back gracefully if not installed.
    
    Note:
        This export is designed for editing and re-uploading via upload_roster().
        ID columns enable accurate user/event matching even if names change.
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    roster = Roster.query.get(roster_id)
    if not roster:
        flash("Roster not found")
        return redirect(url_for('rosters.index'))

    if pd is None or openpyxl is None:
        flash("Excel functionality not available. Please install pandas and openpyxl.")
        return redirect(url_for('rosters.view_roster', roster_id=roster_id))

    # Get competitors and judges from the roster
    competitors = Roster_Competitors.query.filter_by(roster_id=roster_id).all()
    judges = Roster_Judge.query.filter_by(roster_id=roster_id).all()

    # Get partner information
    from mason_snd.models.rosters import Roster_Partners
    roster_partners = Roster_Partners.query.filter_by(roster_id=roster_id).all()
    
    # Build partnership map for quick lookup
    partnership_map = {}
    for partnership in roster_partners:
        partnership_map[partnership.partner1_user_id] = partnership.partner2_user_id
        partnership_map[partnership.partner2_user_id] = partnership.partner1_user_id

    # Build users and events dicts
    user_ids = set([comp.user_id for comp in competitors])
    event_ids = set([comp.event_id for comp in competitors])
    users = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()}
    events = {e.id: e for e in Event.query.filter(Event.id.in_(event_ids)).all()}

    # Judges for the roster
    judge_user_ids = [j.user_id for j in judges if j.user_id]
    child_user_ids = [j.child_id for j in judges if j.child_id]
    all_judge_user_ids = list(set(judge_user_ids + child_user_ids))
    judge_users = {u.id: u for u in User.query.filter(User.id.in_(all_judge_user_ids)).all() if all_judge_user_ids}

    # Create Excel file with multiple sheets
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # JUDGES SHEET - fully editable
    judges_data = []
    for judge in judges:
        judge_name = f"{judge_users[judge.user_id].first_name} {judge_users[judge.user_id].last_name}" if judge.user_id in judge_users else 'Unknown'
        child_name = f"{judge.child.first_name} {judge.child.last_name}" if judge.child else ''
        event_name = judge.event.event_name if judge.event else 'Unknown'
        event_type = 'Unknown'
        if judge.event:
            if judge.event.event_type == 0:
                event_type = 'Speech'
            elif judge.event.event_type == 1:
                event_type = 'LD'
            elif judge.event.event_type == 2:
                event_type = 'PF'
        
        judges_data.append({
            'Judge Name': judge_name,
            'Child': child_name,
            'Event': event_name,
            'Category': event_type,
            'Number People Bringing': judge.people_bringing or 0,
            'Judge ID': judge.user_id,
            'Child ID': judge.child_id if judge.child_id else '',
            'Event ID': judge.event_id if judge.event_id else ''
        })
    
    judges_df = pd.DataFrame(judges_data)
    judges_df.to_excel(writer, sheet_name='Judges', index=False)

    # RANK VIEW SHEET - primary sheet for competitor editing
    # Build rank_view with actual ranking
    rank_view = []
    event_competitors_dict = {}
    
    for comp in competitors:
        eid = comp.event_id
        if eid not in event_competitors_dict:
            event_competitors_dict[eid] = []
        event_competitors_dict[eid].append(comp)
    
    # Rank competitors within each event by weighted points
    for event_id, comps in event_competitors_dict.items():
        sorted_comps = sorted(
            comps, 
            key=lambda c: calculate_weighted_points(users.get(c.user_id)),
            reverse=True
        )
        for rank, comp in enumerate(sorted_comps, start=1):
            rank_view.append({
                'comp': comp,
                'rank': rank,
                'event_id': event_id
            })
    
    rank_data = []
    for item in rank_view:
        comp = item['comp']
        user = users.get(comp.user_id)
        event = events.get(comp.event_id)
        
        user_name = f"{user.first_name} {user.last_name}" if user else 'Unknown'
        event_name = event.event_name if event else 'Unknown Event'
        event_type = 'Unknown'
        if event:
            if event.event_type == 0:
                event_type = 'Speech'
            elif event.event_type == 1:
                event_type = 'LD'
            elif event.event_type == 2:
                event_type = 'PF'
        
        # Get partner information
        partner_name = ''
        partner_id = ''
        if comp.user_id in partnership_map:
            partner_id = partnership_map[comp.user_id]
            if partner_id in users:
                partner_name = f"{users[partner_id].first_name} {users[partner_id].last_name}"
        
        rank_data.append({
            'Rank': item['rank'],
            'Competitor Name': user_name,
            'Partner': partner_name,
            'Weighted Points': calculate_weighted_points(user),
            'Event': event_name,
            'Category': event_type,
            'Status': 'Active',  # Could be extended to show more statuses
            'User ID': comp.user_id,
            'Partner ID': partner_id if partner_id else '',
            'Event ID': comp.event_id
        })
    
    rank_df = pd.DataFrame(rank_data)
    rank_df.to_excel(writer, sheet_name='Rank View', index=False)

    # EVENT VIEW SHEETS - one for each event
    for event_id, comps in event_competitors_dict.items():
        event = events.get(event_id)
        event_name = event.event_name if event else f'Event {event_id}'
        event_type = 'Unknown'
        if event:
            if event.event_type == 0:
                event_type = 'Speech'
            elif event.event_type == 1:
                event_type = 'LD'
            elif event.event_type == 2:
                event_type = 'PF'
        
        # Sort by weighted points for ranking
        sorted_comps = sorted(
            comps,
            key=lambda c: calculate_weighted_points(users.get(c.user_id)),
            reverse=True
        )
        
        event_data = []
        for rank, comp in enumerate(sorted_comps, start=1):
            user = users.get(comp.user_id)
            user_name = f"{user.first_name} {user.last_name}" if user else 'Unknown'
            
            # Get partner information
            partner_name = ''
            partner_id = ''
            if comp.user_id in partnership_map:
                partner_id = partnership_map[comp.user_id]
                if partner_id in users:
                    partner_name = f"{users[partner_id].first_name} {users[partner_id].last_name}"
            
            event_data.append({
                'Event': event_name,
                'Category': event_type,
                'Rank': rank,
                'Competitor': user_name,
                'Partner': partner_name,
                'Weighted Points': calculate_weighted_points(user),
                'User ID': comp.user_id,
                'Partner ID': partner_id if partner_id else '',
                'Event ID': comp.event_id
            })
        
        event_df = pd.DataFrame(event_data)
        # Limit sheet name length and remove invalid characters
        sheet_name = event_name[:30].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace(':', '-').replace('[', '-').replace(']', '-')
        event_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Add formatting to make it clear what can be edited
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Format all sheets
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # ID columns (read-only indicators) - lighter background
        id_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        
        # Editable columns - white/normal
        editable_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply column-specific formatting
        for col_idx, column in enumerate(worksheet.iter_cols(min_row=2), start=1):
            col_letter = get_column_letter(col_idx)
            header_value = worksheet[f'{col_letter}1'].value
            
            # Adjust column width
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Color code columns
            if header_value and 'ID' in str(header_value):
                # ID columns - indicate these are used for matching, but names take precedence for display
                for cell in column:
                    cell.fill = id_fill
            else:
                # Editable columns
                for cell in column:
                    if header_value in ['Rank', 'Weighted Points', 'Status', 'Category', 'Event']:
                        # These are informational/calculated - light gray
                        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    else:
                        # Names and other editable fields - white
                        cell.fill = editable_fill

    writer.close()
    output.seek(0)

    filename = f"roster_{roster.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, 
                     as_attachment=True, 
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@rosters_bp.route('/rename_roster/<int:roster_id>', methods=['GET', 'POST'])
@prevent_race_condition('rename_roster', min_interval=1.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('rosters.view_roster', roster_id=request.view_args.get('roster_id'))))
def rename_roster(roster_id):
    """Change roster name for better organization.
    
    Methods:
        GET: Display rename form with current name.
        POST: Update roster name in database.
    
    URL Parameters:
        roster_id (int): Roster primary key.
    
    Form Fields (POST):
        - new_name (str): New roster name.
    
    Race Condition Protection:
        @prevent_race_condition decorator prevents duplicate submissions within 1 second.
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        GET: rosters/rename_roster.html with current roster.
        POST: Redirects to view_roster with success message.
    
    Use Cases:
        - Clarify roster purpose (e.g., "Harvard Final Roster v2")
        - Add notes (e.g., "Revised after drops")
        - Standardize naming convention
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    roster = Roster.query.get(roster_id)
    if not roster:
        flash("Roster not found")
        return redirect(url_for('rosters.index'))

    if request.method == 'POST':
        new_name = request.form.get('new_name')
        if new_name:
            roster.name = new_name
            db.session.commit()
            flash(f"Roster renamed to '{new_name}'")
        return redirect(url_for('rosters.view_roster', roster_id=roster_id))

    return render_template('rosters/rename_roster.html', roster=roster)

@rosters_bp.route('/delete_roster/<int:roster_id>')
def delete_roster(roster_id):
    """Permanently delete a roster and all associated data.
    
    Deletion Process:
    1. Delete all Roster_Competitors entries
    2. Delete all Roster_Judge entries
    3. Delete Roster record itself
    
    URL Parameters:
        roster_id (int): Roster primary key.
    
    Database Records Deleted:
        - Roster_Competitors: All competitor assignments
        - Roster_Judge: All judge assignments
        - Roster: Main roster record
    
    Cascade Deletions:
        Foreign key constraints may also delete:
        - User_Published_Rosters (if roster was published)
        - Roster_Penalty_Entries (penalty tracking)
        - Roster_Partners (partner relationships)
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        Redirects to rosters.index with success message.
        Flash error if roster not found.
    
    Warning:
        This is PERMANENT deletion. Cannot be undone.
        Consider unpublishing instead if you may need roster later.
    
    Note:
        Deleting a published roster removes it from all user profiles
        (User_Published_Rosters entries are cascaded).
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    roster = Roster.query.get(roster_id)
    if not roster:
        flash("Roster not found")
        return redirect(url_for('rosters.index'))

    # Delete associated competitors, judges, and partners
    Roster_Competitors.query.filter_by(roster_id=roster_id).delete()
    Roster_Judge.query.filter_by(roster_id=roster_id).delete()
    
    from mason_snd.models.rosters import Roster_Partners
    Roster_Partners.query.filter_by(roster_id=roster_id).delete()
    
    # Delete the roster itself
    db.session.delete(roster)
    db.session.commit()
    
    flash("Roster deleted successfully")
    return redirect(url_for('rosters.index'))

@rosters_bp.route('/upload_roster', methods=['GET', 'POST'])
@prevent_race_condition('upload_roster', min_interval=2.0, redirect_on_duplicate=lambda uid, form: redirect(url_for('rosters.index')))
def upload_roster():
    """Import Excel file to create or update roster with smart name reconciliation.
    
    Methods:
        GET: Display upload form with roster list (for updating existing rosters).
        POST: Process Excel file and create/update roster.
    
    Form Fields (POST):
        - file: Excel file (.xlsx) from download_roster() or similar format
        - roster_name: Name for new roster (optional, default: "Uploaded Roster {timestamp}")
        - roster_id: ID of existing roster to update (optional, for updates)
    
    Smart Reconciliation Algorithm:
        User Matching (Priority Order):
            1. User ID: If valid ID in 'User ID' column, use it (name changes ignored)
            2. Exact Name Match: "First Last" → User.query.filter_by(first_name, last_name)
            3. Fuzzy Name Match: Case-insensitive matching
            4. No Match: Log warning, skip entry
        
        Event Matching (Priority Order):
            1. Event ID: If valid ID in 'Event ID' column, use it
            2. Event Name: Event.query.filter_by(event_name)
            3. No Match: Log warning, skip entry
    
    Excel Format Expected:
        Required Sheets:
            - Judges: Judge Name, Child, Event, Category, Number People Bringing,
                     Judge ID*, Child ID*, Event ID*
            - Rank View: Competitor Name, Event, User ID*, Event ID*
                        (Partner, Weighted Points, Rank, Category are optional)
        
        Optional Sheets:
            - Event-specific sheets: Same columns as Rank View, event-scoped
        
        * ID columns are optional but enable accurate reconciliation
    
    Update vs Create:
        - If roster_id provided: Updates existing roster (clears old data)
        - Otherwise: Creates new roster with provided name
        - Update clears: Roster_Competitors, Roster_Judge, Roster_Partners
    
    Change Tracking:
        Returns log with:
        - judges: List of added judges with row numbers
        - competitors: List of added competitors with events
        - warnings: List of unmatched entries (up to 5 shown in flash)
    
    Race Condition Protection:
        @prevent_race_condition (2 second interval) prevents duplicate uploads.
    
    Access Control:
        Requires role >= 2 (admin). Non-admins redirected to main.index.
    
    Returns:
        GET: rosters/upload_roster.html with roster list.
        POST: Redirects to view_roster with change summary.
        Error: Flash error message, redirect to upload form.
    
    Dependencies:
        - pandas: Excel reading
        - openpyxl: Excel file format support
        Falls back with error message if not installed.
    
    Note:
        Smart reconciliation allows editing names in Excel while maintaining
        correct user associations via ID columns. This enables collaborative
        editing workflows and roster corrections.
    """
    user_id = session.get('user_id')
    user = User.query.filter_by(id=user_id).first()

    if not user_id:
        flash("Log In First")
        return redirect_to_login()

    if user.role < 2:
        flash("You are not authorized to access this page")
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash("No file selected")
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash("No file selected")
            return redirect(request.url)
        
        if not file.filename.endswith('.xlsx'):
            flash("Please upload an Excel (.xlsx) file")
            return redirect(request.url)

        if pd is None or openpyxl is None:
            flash("Excel functionality not available. Please install pandas and openpyxl.")
            return redirect(url_for('rosters.index'))

        try:
            # Read the Excel file
            excel_file = pd.ExcelFile(file)
            
            # Get roster name and ID from form
            roster_name = request.form.get('roster_name', f"Uploaded Roster {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            roster_id = request.form.get('roster_id')  # If updating existing roster
            
            # Determine if we're updating or creating
            if roster_id:
                new_roster = Roster.query.get(int(roster_id))
                if not new_roster:
                    flash("Roster not found")
                    return redirect(url_for('rosters.index'))
                
                # Clear existing data
                Roster_Competitors.query.filter_by(roster_id=new_roster.id).delete()
                Roster_Judge.query.filter_by(roster_id=new_roster.id).delete()
                from mason_snd.models.rosters import Roster_Partners
                Roster_Partners.query.filter_by(roster_id=new_roster.id).delete()
            else:
                # Create new roster
                tz = pytz.timezone('US/Eastern')
                new_roster = Roster(name=roster_name, date_made=datetime.now(tz))
                db.session.add(new_roster)
                db.session.flush()  # Get the ID

            # Helper function to find user by ID or name with smart matching
            def find_user_smart(user_id_val, name_val):
                """
                Smart user finder:
                1. If User ID is provided and valid, use it (name changes are ignored)
                2. Otherwise, try to match by exact name
                3. Fall back to fuzzy name matching
                """
                # Try User ID first (most reliable)
                if pd.notna(user_id_val):
                    try:
                        user = User.query.get(int(user_id_val))
                        if user:
                            return user
                    except (ValueError, TypeError):
                        pass
                
                # Try exact name match
                if name_val and str(name_val).strip():
                    name_parts = str(name_val).strip().split()
                    if len(name_parts) >= 2:
                        first_name = name_parts[0]
                        last_name = ' '.join(name_parts[1:])
                        user = User.query.filter_by(first_name=first_name, last_name=last_name).first()
                        if user:
                            return user
                        
                        # Try fuzzy matching (case-insensitive)
                        user = User.query.filter(
                            db.func.lower(User.first_name) == first_name.lower(),
                            db.func.lower(User.last_name) == last_name.lower()
                        ).first()
                        if user:
                            return user
                
                return None

            # Process judges sheet with smart reconciliation
            changes_log = {'judges': [], 'competitors': [], 'warnings': []}
            
            if 'Judges' in excel_file.sheet_names:
                judges_df = pd.read_excel(file, sheet_name='Judges')
                for idx, row in judges_df.iterrows():
                    # Find judge user with smart matching
                    judge_user = find_user_smart(
                        row.get('Judge ID'),
                        row.get('Judge Name')
                    )
                    
                    # Find child user with smart matching
                    child_user = find_user_smart(
                        row.get('Child ID'),
                        row.get('Child', '')
                    )
                    
                    # Find event by ID (prioritize ID over name)
                    event = None
                    if 'Event ID' in row and pd.notna(row['Event ID']):
                        event = Event.query.get(int(row['Event ID']))
                    elif 'Event' in row and pd.notna(row['Event']):
                        event = Event.query.filter_by(event_name=str(row['Event'])).first()
                    
                    # Get people_bringing
                    people_bringing = 0
                    if 'Number People Bringing' in row and pd.notna(row['Number People Bringing']):
                        people_bringing = int(row['Number People Bringing'])
                    elif event:
                        if event.event_type == 0:
                            people_bringing = 6
                        elif event.event_type == 1:
                            people_bringing = 2
                        elif event.event_type == 2:
                            people_bringing = 4
                    
                    if judge_user:
                        rj = Roster_Judge(
                            user_id=judge_user.id,
                            child_id=child_user.id if child_user else None,
                            event_id=event.id if event else None,
                            roster_id=new_roster.id,
                            people_bringing=people_bringing
                        )
                        db.session.add(rj)
                        changes_log['judges'].append(f"Row {idx+2}: Added judge {judge_user.first_name} {judge_user.last_name}")
                    else:
                        changes_log['warnings'].append(f"Row {idx+2} in Judges: Could not find user '{row.get('Judge Name', 'Unknown')}'")

            # Process competitors from Rank View sheet with smart reconciliation
            if 'Rank View' in excel_file.sheet_names:
                rank_df = pd.read_excel(file, sheet_name='Rank View')
                
                # Sort by rank if available to preserve order
                if 'Rank' in rank_df.columns:
                    rank_df = rank_df.sort_values('Rank')
                
                for idx, row in rank_df.iterrows():
                    # Find user with smart matching
                    user = find_user_smart(
                        row.get('User ID'),
                        row.get('Competitor Name')
                    )
                    
                    # Find event by ID (prioritize ID)
                    event = None
                    if 'Event ID' in row and pd.notna(row['Event ID']):
                        event = Event.query.get(int(row['Event ID']))
                    elif 'Event' in row and pd.notna(row['Event']):
                        event = Event.query.filter_by(event_name=str(row['Event'])).first()
                    
                    if user and event:
                        rc = Roster_Competitors(
                            user_id=user.id,
                            event_id=event.id,
                            judge_id=None,
                            roster_id=new_roster.id
                        )
                        db.session.add(rc)
                        changes_log['competitors'].append(f"Row {idx+2}: Added {user.first_name} {user.last_name} to {event.event_name}")
                    else:
                        warning_msg = f"Row {idx+2} in Rank View: "
                        if not user:
                            warning_msg += f"Could not find user '{row.get('Competitor Name', 'Unknown')}'"
                        if not event:
                            warning_msg += f"Could not find event '{row.get('Event', 'Unknown')}'"
                        changes_log['warnings'].append(warning_msg)

            # Process event view sheets (for additional competitor data)
            for sheet_name in excel_file.sheet_names:
                if sheet_name not in ['Judges', 'Rank View']:
                    event_df = pd.read_excel(file, sheet_name=sheet_name)
                    
                    # Try to find the event for this sheet
                    event = None
                    if 'Event ID' in event_df.columns and len(event_df) > 0:
                        event_id = event_df.iloc[0]['Event ID']
                        if pd.notna(event_id):
                            event = Event.query.get(int(event_id))
                    
                    if not event and 'Event' in event_df.columns and len(event_df) > 0:
                        event_name = event_df.iloc[0]['Event']
                        if pd.notna(event_name):
                            event = Event.query.filter_by(event_name=str(event_name)).first()
                    
                    # Process competitors in this event sheet
                    for idx, row in event_df.iterrows():
                        user = find_user_smart(
                            row.get('User ID'),
                            row.get('Competitor')
                        )
                        
                        if user and event:
                            # Check if this competitor is already added
                            existing = Roster_Competitors.query.filter_by(
                                roster_id=new_roster.id,
                                user_id=user.id,
                                event_id=event.id
                            ).first()
                            
                            if not existing:
                                rc = Roster_Competitors(
                                    user_id=user.id,
                                    event_id=event.id,
                                    judge_id=None,
                                    roster_id=new_roster.id
                                )
                                db.session.add(rc)

            # Auto-fill roster with additional competitors to use all judge capacity
            db.session.flush()  # Ensure judges and existing competitors are saved
            _auto_fill_roster_from_signups(new_roster.id)
            _redistribute_competitors_for_roster(new_roster.id)
            
            db.session.commit()
            
            # Show success message with changes summary
            success_msg = f"Roster '{new_roster.name}' {'updated' if roster_id else 'created'} successfully! "
            success_msg += f"Added {len(changes_log['competitors'])} competitors and {len(changes_log['judges'])} judges."
            if changes_log['warnings']:
                success_msg += f" {len(changes_log['warnings'])} warnings."
            
            flash(success_msg, 'success')
            
            # Show warnings if any
            for warning in changes_log['warnings'][:5]:  # Limit to first 5 warnings
                flash(warning, 'warning')
            
            return redirect(url_for('rosters.view_roster', roster_id=new_roster.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Error processing file: {str(e)}", 'error')
            import traceback
            print(traceback.format_exc())
            return redirect(request.url)

    # GET request - show upload form with list of rosters
    rosters = Roster.query.order_by(Roster.date_made.desc()).all()
    return render_template('rosters/upload_roster.html', rosters=rosters)
